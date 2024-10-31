import os
import openpyxl
from openpyxl import Workbook
import math
import time
import copy

OUTPUT_FILE = 'VRPTW_LuisaMariaAlvarez_VND_5.xlsx'
INSTANCES_DIR = 'instances'  

class Node:
    def __init__(self, id, x, y, demand, early, late, service_time):
        self.id = id
        self.x = x
        self.y = y
        self.demand = demand
        self.early = early
        self.late = late
        self.service_time = service_time

class Vehicle:
    def __init__(self, capacity):
        self.capacity = capacity
        self.route = []
        self.load = 0
        self.time = 0
        self.arrival_times = []

class VRPTW:
    def __init__(self, nodes, vehicle_capacity):
        self.nodes = nodes
        self.vehicle_capacity = vehicle_capacity
        self.depot = nodes[0]
        self.vehicles = []

    def distance(self, node1, node2):
        return round(math.sqrt((node1.x - node2.x)**2 + (node1.y - node2.y)**2), 3)

    def calculate_total_distance(self):
        total_distance = 0
        for vehicle in self.vehicles:
            route = vehicle.route
            for i in range(len(route) - 1):
                total_distance += self.distance(route[i], route[i + 1])
        return round(total_distance, 3)

    def is_feasible(self, vehicle, node):
        if vehicle.load + node.demand > vehicle.capacity:
            return False
        
        arrival_time = vehicle.time + self.distance(vehicle.route[-1], node)
        
        if arrival_time > node.late:
            return False
        return_time_to_depot = max(arrival_time, node.early) + node.service_time + self.distance(node, self.depot)

        if return_time_to_depot > self.depot.late:
            return False
        
        return True

    def update_vehicle(self, vehicle, node):
        arrival_time = vehicle.time + self.distance(vehicle.route[-1], node)
        service_start = max(arrival_time, node.early)
        vehicle.time = service_start + node.service_time
        
        vehicle.route.append(node)
        vehicle.load += node.demand
        vehicle.arrival_times.append(round(arrival_time, 3))

    def construct_solution(self):
        unassigned = self.nodes[1:]
        max_attempts = len(self.nodes) * 2
        attempts = 0
        
        while unassigned and attempts < max_attempts:
            vehicle = Vehicle(self.vehicle_capacity)
            vehicle.route = [self.depot]
            vehicle.time = 0
            vehicle.arrival_times = [0]
            assigned_any = False
            
            while True:
                feasible_nodes = [node for node in unassigned if self.is_feasible(vehicle, node)]
                
                if not feasible_nodes:
                    break
                
                next_node = min(feasible_nodes, key=lambda node: self.distance(vehicle.route[-1], node))
                
                self.update_vehicle(vehicle, next_node)
                unassigned.remove(next_node)
                assigned_any = True
            
            vehicle.route.append(self.depot)
            vehicle.arrival_times.append(round(vehicle.time + self.distance(vehicle.route[-2], self.depot), 3))
            self.vehicles.append(vehicle)
            
            if not assigned_any:
                attempts += 1
            else:
                attempts = 0
        
        return self.vehicles

    def two_opt_best_improvement(self):
        improved = True
        while improved:
            improved = False
            best_improvement = 0
            best_swap = None

            for v_index, vehicle in enumerate(self.vehicles):
                route = vehicle.route
                for i in range(1, len(route) - 2):
                    for j in range(i + 1, len(route) - 1):
                        new_route = route[:i] + route[i:j+1][::-1] + route[j+1:]
                        if self.is_route_feasible(new_route):
                            old_distance = self.route_distance(route)
                            new_distance = self.route_distance(new_route)
                            improvement = old_distance - new_distance
                            if improvement > best_improvement:
                                best_improvement = improvement
                                best_swap = (v_index, i, j)

            if best_swap:
                v_index, i, j = best_swap
                vehicle = self.vehicles[v_index]
                vehicle.route = vehicle.route[:i] + vehicle.route[i:j+1][::-1] + vehicle.route[j+1:]
                self.update_vehicle_times(vehicle)
                improved = True

    def segment_exchange_best_improvement(self):
        improved = True
        while improved:
            improved = False
            best_improvement = 0
            best_exchange = None

            for i in range(len(self.vehicles)):
                for j in range(i + 1, len(self.vehicles)):
                    exchange_result = self.find_best_exchange(self.vehicles[i], self.vehicles[j])
                    if exchange_result:
                        improvement, exchange_info = exchange_result
                        if improvement > best_improvement:
                            best_improvement = improvement
                            best_exchange = (i, j, exchange_info)

            if best_exchange:
                i, j, (i_start, j_start, segment_length) = best_exchange
                self.apply_exchange(self.vehicles[i], self.vehicles[j], i_start, j_start, segment_length)
                improved = True

    def apply_exchange(self, vehicle1, vehicle2, i, j, segment_length):
        new_route1 = vehicle1.route[:i] + vehicle2.route[j:j+segment_length] + vehicle1.route[i+segment_length:]
        new_route2 = vehicle2.route[:j] + vehicle1.route[i:i+segment_length] + vehicle2.route[j+segment_length:]

        vehicle1.route = new_route1
        vehicle2.route = new_route2
        self.update_vehicle_times(vehicle1)
        self.update_vehicle_times(vehicle2)

    def find_best_exchange(self, vehicle1, vehicle2):
        best_improvement = 0
        best_exchange = None

        for i in range(1, len(vehicle1.route) - 2):
            for j in range(1, len(vehicle2.route) - 2):
                for segment_length in range(1, min(len(vehicle1.route) - i - 1, len(vehicle2.route) - j - 1)):
                    new_route1 = vehicle1.route[:i] + vehicle2.route[j:j+segment_length] + vehicle1.route[i+segment_length:]
                    new_route2 = vehicle2.route[:j] + vehicle1.route[i:i+segment_length] + vehicle2.route[j+segment_length:]

                    if (self.is_route_feasible(new_route1) and self.is_route_feasible(new_route2)):
                        old_distance = self.route_distance(vehicle1.route) + self.route_distance(vehicle2.route)
                        new_distance = self.route_distance(new_route1) + self.route_distance(new_route2)
                        improvement = old_distance - new_distance

                        if improvement > best_improvement:
                            best_improvement = improvement
                            best_exchange = (i, j, segment_length)

        return (best_improvement, best_exchange) if best_exchange else None


    def relocation_best_improvement(self):
        improved = True
        while improved:
            improved = False
            best_improvement = 0
            best_move = None

            for v1_index, vehicle1 in enumerate(self.vehicles):
                for i in range(1, len(vehicle1.route) - 1): 
                    node = vehicle1.route[i]
                    for j in range(1, len(vehicle1.route) - 1):
                        if i != j and i != j - 1:
                            new_route = vehicle1.route[:i] + vehicle1.route[i+1:]
                            new_route = new_route[:j] + [node] + new_route[j:]
                            if self.is_route_feasible(new_route):
                                improvement = self.route_distance(vehicle1.route) - self.route_distance(new_route)
                                if improvement > best_improvement:
                                    best_improvement = improvement
                                    best_move = (v1_index, v1_index, i, j)

                    for v2_index, vehicle2 in enumerate(self.vehicles):
                        if v1_index != v2_index:
                            for j in range(1, len(vehicle2.route)):
                                new_route1 = vehicle1.route[:i] + vehicle1.route[i+1:]
                                new_route2 = vehicle2.route[:j] + [node] + vehicle2.route[j:]
                                if self.is_route_feasible(new_route1) and self.is_route_feasible(new_route2):
                                    improvement = (self.route_distance(vehicle1.route) + self.route_distance(vehicle2.route)) - \
                                                  (self.route_distance(new_route1) + self.route_distance(new_route2))
                                    if improvement > best_improvement:
                                        best_improvement = improvement
                                        best_move = (v1_index, v2_index, i, j)

            if best_move:
                v1_index, v2_index, i, j = best_move
                node = self.vehicles[v1_index].route[i]
                
                del self.vehicles[v1_index].route[i]
                self.vehicles[v1_index].load -= node.demand
                
                self.vehicles[v2_index].route.insert(j, node)
                self.vehicles[v2_index].load += node.demand
                
                self.update_vehicle_times(self.vehicles[v1_index])
                self.update_vehicle_times(self.vehicles[v2_index])
                
                improved = True

    def is_route_feasible(self, route):
        time = 0
        load = 0
        
        for i in range(len(route) - 1):
            current_node = route[i]
            next_node = route[i + 1]
            
            arrival_time = time + self.distance(current_node, next_node)
            if arrival_time > next_node.late:
                return False
            
            service_start = max(arrival_time, next_node.early)
            time = service_start + next_node.service_time
            
            load += current_node.demand
            if load > self.vehicle_capacity:
                return False
        
        return time <= self.depot.late

    def route_distance(self, route):
        return sum(self.distance(route[i], route[i+1]) for i in range(len(route) - 1))

    def update_vehicle_times(self, vehicle):
        time = 0
        vehicle.arrival_times = [0]
        vehicle.load = 0
        
        for i in range(1, len(vehicle.route)):
            prev_node = vehicle.route[i-1]
            curr_node = vehicle.route[i]
            
            arrival_time = time + self.distance(prev_node, curr_node)
            service_start = max(arrival_time, curr_node.early)
            time = service_start + curr_node.service_time
            
            vehicle.arrival_times.append(round(arrival_time, 3))
            vehicle.load += curr_node.demand
        
        vehicle.time = time

def read_instance(filename):
    with open(filename, 'r') as f:
        n, capacity = map(int, f.readline().split())
        nodes = []
        for _ in range(n + 1):
            id, x, y, demand, early, late, service_time = map(float, f.readline().split())
            nodes.append(Node(int(id), x, y, demand, early, late, service_time))
    return nodes, capacity

def save_results_to_excel(instance_name, vehicles, total_distance, computation_time, vehicle_capacity):
    if not os.path.exists(OUTPUT_FILE):
        workbook = Workbook()
        workbook.remove(workbook.active) 
    else:
        workbook = openpyxl.load_workbook(OUTPUT_FILE)

    sheet = workbook.create_sheet(title=instance_name)
    
    used_vehicles = [vehicle for vehicle in vehicles if len(vehicle.route) > 2]

    sheet.append([len(used_vehicles), round(total_distance, 3), round(computation_time, 3)])

    for vehicle in used_vehicles:
        route = [node.id for node in vehicle.route]
        adjusted_times = [max(at, node.early) for at, node in zip(vehicle.arrival_times, vehicle.route)]

        route_info = [len(route) - 2] + route + adjusted_times + [vehicle.load]
        sheet.append(route_info)

    workbook.save(OUTPUT_FILE)


class VND:
    def __init__(self, vrptw, time_limit):
        self.vrptw = vrptw
        self.time_limit = time_limit
        self.start_time = None
        self.best_solution = None
        self.best_distance = float('inf')
        
    def get_elapsed_time(self):
        return time.time() - self.start_time
        
    def check_time_limit(self):
        return self.get_elapsed_time() < self.time_limit

    def apply_neighborhood(self, k):
        if k == 0:
            self.vrptw.relocation_best_improvement()
        elif k == 1:
            self.vrptw.two_opt_best_improvement()
        elif k == 2:
            self.vrptw.segment_exchange_best_improvement()

    def save_best_solution(self):
        self.best_solution = copy.deepcopy(self.vrptw.vehicles)
        self.best_distance = self.vrptw.calculate_total_distance()

    def restore_best_solution(self):
        self.vrptw.vehicles = copy.deepcopy(self.best_solution)

    def solve(self):
        self.start_time = time.time()
        k_max = 5

        self.vrptw.construct_solution()
        self.save_best_solution()
        
        while self.check_time_limit():
            k = 0  
            while k < k_max and self.check_time_limit():
                current_distance = self.vrptw.calculate_total_distance()
                self.apply_neighborhood(k)
                new_distance = self.vrptw.calculate_total_distance()
                
                if new_distance < current_distance:
                    self.save_best_solution()
                    k = 0
                else:
                    k += 1
                    
        self.restore_best_solution()
        return self.best_distance, self.get_elapsed_time() * 1000  # Return time in milliseconds

def read_time_limit(instance_name):
    # Read time limit from TimeLimit.xlsx
    wb = openpyxl.load_workbook('TimeLimit.xlsx')
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == instance_name:
            return float(row[1].value)
    return 60  # Default time limit if not found

def solve_instance(instance_path):
    nodes, capacity = read_instance(instance_path)
    vrptw = VRPTW(nodes, capacity)
    
    # Get instance name and time limit
    instance_name = os.path.basename(instance_path).replace('.txt', '')
    time_limit = read_time_limit(instance_name)
    
    # Create and run VND
    vnd = VND(vrptw, time_limit)
    total_distance, computation_time = vnd.solve()
    
    routes = [(vehicle.route, vehicle.arrival_times[-1], vehicle.load) 
              for vehicle in vrptw.vehicles]
    
    return vrptw.vehicles, total_distance, computation_time, routes, capacity

def main():
    for instance_filename in os.listdir(INSTANCES_DIR):
        if instance_filename.endswith('.txt'):
            instance_path = os.path.join(INSTANCES_DIR, instance_filename)
            instance_name = instance_filename.replace('.txt', '')
            
            print(f"\nSolving instance: {instance_name}")
            print(f"Time limit: {read_time_limit(instance_name)} seconds")
            
            vehicles, total_distance, computation_time, routes, vehicle_capacity = \
                solve_instance(instance_path)
            
            print(f"Vehicles used: {len(vehicles)}")
            print(f"Total Distance: {total_distance}")
            print(f"Computation Time: {computation_time:.2f} ms")
            print(f"Vehicle Capacity: {vehicle_capacity}")
            print("Routes:")
            for i, (route, arrival_time, load) in enumerate(routes, 1):
                print(f"  Route {i}:")
                print(f"    Nodes: {' -> '.join(str(node.id) for node in route)}")
                print(f"    Arrival Time: {arrival_time:.2f}")
                print(f"    Load: {load}")
            
            save_results_to_excel(instance_name, vehicles, total_distance, 
                                computation_time, vehicle_capacity)

if __name__ == '__main__':
    main()