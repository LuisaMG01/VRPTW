import os
import openpyxl
from openpyxl import Workbook
import math
import time
import random

OUTPUT_FILE = 'VRPTW_LuisaMariaAlvarez_multistart_els_tabu_time.xlsx'
INSTANCES_DIR = 'instances'  
time_instance = 'TimeLimit.xlsx' 

class Node:
    def __init__(self, id, x, y, demand, early, late, service_time):
        self.id = id
        self.x = x
        self.y = y
        self.demand = demand
        self.early = early
        self.late = late
        self.service_time = service_time

class Vehicle:
    def __init__(self, capacity):
        self.capacity = capacity
        self.route = []
        self.load = 0
        self.time = 0
        self.arrival_times = []
        
    def reset(self):
        self.route = []
        self.load = 0
        self.time = 0
        self.arrival_times = []

class TabuSearch:
    def __init__(self, vrptw, tabu_tenure=10, max_iterations=100):
        self.vrptw = vrptw
        self.tabu_tenure = tabu_tenure
        self.max_iterations = max_iterations
        self.tabu_list = []  
        self.best_solution = None
        self.best_distance = float('inf')
        
    def is_tabu(self, move):
        return move in self.tabu_list
    
    def update_tabu_list(self, move):
        self.tabu_list.append(move)
        if len(self.tabu_list) > self.tabu_tenure:
            self.tabu_list.pop(0)
    
    def get_neighborhood_moves(self):
        moves = []
        for i, vehicle1 in enumerate(self.vrptw.vehicles):
            for pos1 in range(1, len(vehicle1.route) - 1):
                node1 = vehicle1.route[pos1]
                
                for pos2 in range(1, len(vehicle1.route) - 1):
                    if pos1 != pos2:
                        moves.append(('swap', i, i, pos1, pos2))
                
                for j, vehicle2 in enumerate(self.vrptw.vehicles):
                    if i != j:
                        for pos2 in range(1, len(vehicle2.route)):
                            moves.append(('relocate', i, j, pos1, pos2))
                            
        return moves
    
    def apply_move(self, move):
        move_type, v1_idx, v2_idx, pos1, pos2 = move
        vehicle1 = self.vrptw.vehicles[v1_idx]
        vehicle2 = self.vrptw.vehicles[v2_idx]
        
        if move_type == 'swap':
            vehicle1.route[pos1], vehicle1.route[pos2] = vehicle1.route[pos2], vehicle1.route[pos1]
            self.vrptw.update_vehicle_times(vehicle1)
            
        elif move_type == 'relocate':
            node = vehicle1.route.pop(pos1)
            vehicle2.route.insert(pos2, node)
            self.vrptw.update_vehicle_times(vehicle1)
            self.vrptw.update_vehicle_times(vehicle2)
    
    def evaluate_move(self, move):
        original_vehicles = [Vehicle(v.capacity) for v in self.vrptw.vehicles]
        for i, v in enumerate(self.vrptw.vehicles):
            original_vehicles[i].route = v.route.copy()
            original_vehicles[i].load = v.load
            original_vehicles[i].time = v.time
            original_vehicles[i].arrival_times = v.arrival_times.copy()
        
        self.apply_move(move)
        
        is_feasible = all(self.vrptw.is_route_feasible(v.route) for v in self.vrptw.vehicles)
        distance = self.vrptw.calculate_total_distance() if is_feasible else float('inf')
        
        self.vrptw.vehicles = original_vehicles
        
        return distance if is_feasible else float('inf')
    
    def run(self):
        current_distance = self.vrptw.calculate_total_distance()
        self.best_distance = current_distance
        self.best_solution = [Vehicle(v.capacity) for v in self.vrptw.vehicles]
        for i, v in enumerate(self.vrptw.vehicles):
            self.best_solution[i].route = v.route.copy()
            self.best_solution[i].load = v.load
            self.best_solution[i].time = v.time
            self.best_solution[i].arrival_times = v.arrival_times.copy()
        
        iterations_without_improvement = 0
        
        for iteration in range(self.max_iterations):
            moves = self.get_neighborhood_moves()
            best_move = None
            best_move_distance = float('inf')
            
            for move in moves:
                if not self.is_tabu(move):
                    move_distance = self.evaluate_move(move)
                    
                    if move_distance < self.best_distance:
                        best_move = move
                        best_move_distance = move_distance
                        break
                    
                    if move_distance < best_move_distance:
                        best_move = move
                        best_move_distance = move_distance
            
            if best_move is None:
                break
            
            self.apply_move(best_move)
            self.update_tabu_list(best_move)
            current_distance = self.vrptw.calculate_total_distance()
            
            if current_distance < self.best_distance:
                self.best_distance = current_distance
                self.best_solution = [Vehicle(v.capacity) for v in self.vrptw.vehicles]
                for i, v in enumerate(self.vrptw.vehicles):
                    self.best_solution[i].route = v.route.copy()
                    self.best_solution[i].load = v.load
                    self.best_solution[i].time = v.time
                    self.best_solution[i].arrival_times = v.arrival_times.copy()
                iterations_without_improvement = 0
            else:
                iterations_without_improvement += 1
            
            if iterations_without_improvement >= 20:
                break
        
        self.vrptw.vehicles = self.best_solution
        return self.best_distance


class VRPTW:
    def __init__(self, nodes, vehicle_capacity):
        self.nodes = nodes
        self.vehicle_capacity = vehicle_capacity
        self.depot = nodes[0]
        self.vehicles = []
        self.best_solution = None
        self.best_distance = float('inf')

    
    def distance(self, node1, node2):
        return round(math.sqrt((node1.x - node2.x)**2 + (node1.y - node2.y)**2), 3)

    def calculate_total_distance(self, vehicles=None):
        if vehicles is None:
            vehicles = self.vehicles
        total_distance = 0
        for vehicle in vehicles:
            route = vehicle.route
            for i in range(len(route) - 1):
                total_distance += self.distance(route[i], route[i + 1])
        return round(total_distance, 3)

    def calculate_arrival_time(self, route, current_time, start_idx):
        for i in range(start_idx, len(route)):
            if i == 0:
                continue
            prev_node = route[i-1]
            current_node = route[i]
            arrival_time = current_time + self.distance(prev_node, current_node)
            current_time = max(arrival_time, current_node.early) + current_node.service_time
        return current_time

    def is_feasible(self, vehicle, node, check_capacity=True):
        if check_capacity and vehicle.load + node.demand > vehicle.capacity:
            return False
        
        arrival_time = vehicle.time + self.distance(vehicle.route[-1], node)
        
        if arrival_time > node.late:
            return False
            
        service_start = max(arrival_time, node.early)
        departure_time = service_start + node.service_time
        
        time_to_depot = departure_time + self.distance(node, self.depot)
        if time_to_depot > self.depot.late:
            return False
        
        return True

    def update_vehicle(self, vehicle, node):
        arrival_time = vehicle.time + self.distance(vehicle.route[-1], node)
        service_start = max(arrival_time, node.early)
        vehicle.time = service_start + node.service_time
        vehicle.route.append(node)
        vehicle.load += node.demand
        vehicle.arrival_times.append(round(arrival_time, 3))

    def try_insert_node(self, vehicle, node, position):
        if position == 0 or position > len(vehicle.route):
            return False
            
        temp_route = vehicle.route.copy()
        temp_route.insert(position, node)
        
        total_load = sum(n.demand for n in temp_route if n != self.depot)
        if total_load > vehicle.capacity:
            return False
            
        current_time = 0
        for i in range(1, len(temp_route)):
            arrival_time = current_time + self.distance(temp_route[i-1], temp_route[i])
            if arrival_time > temp_route[i].late:
                return False
            current_time = max(arrival_time, temp_route[i].early) + temp_route[i].service_time
            
        return True
    
    def improve_route_distance(self, vehicle):
        route = vehicle.route
        best_distance = self.route_distance(route)
        improved = True

        while improved:
            improved = False
            for i in range(1, len(route) - 2):
                for j in range(i + 1, len(route) - 1):
                    new_route = route[:i] + route[i:j + 1][::-1] + route[j + 1:]
                    new_distance = self.route_distance(new_route)
                    
                    if new_distance < best_distance and self.is_route_feasible(new_route):
                        route = new_route
                        best_distance = new_distance
                        improved = True
                        
            vehicle.route = route
            self.update_vehicle_times(vehicle)

    def construct_initial_solution(self):
        unassigned = self.nodes[1:] 
        unassigned.sort(key=lambda x: (x.early, x.late))
        
        while unassigned:
            vehicle = Vehicle(self.vehicle_capacity)
            vehicle.route = [self.depot]
            vehicle.arrival_times = [0]
            
            assigned_nodes = []
            for node in unassigned:
                if self.is_feasible(vehicle, node):
                    self.update_vehicle(vehicle, node)
                    assigned_nodes.append(node)
                    
            for node in assigned_nodes:
                unassigned.remove(node)
            
            if len(vehicle.route) > 1:
                vehicle.route.append(self.depot)
                vehicle.arrival_times.append(
                    round(vehicle.time + self.distance(vehicle.route[-2], self.depot), 3)
                )
                self.vehicles.append(vehicle)
            
            elif unassigned:
                vehicle = Vehicle(self.vehicle_capacity)
                vehicle.route = [self.depot]
                vehicle.arrival_times = [0]
                node = unassigned[0]
                self.update_vehicle(vehicle, node)
                vehicle.route.append(self.depot)
                vehicle.arrival_times.append(
                    round(vehicle.time + self.distance(node, self.depot), 3)
                )
                self.vehicles.append(vehicle)
                unassigned.remove(node)

    def local_search(self, max_iterations=100):
        current_iteration = 0
        best_distance = self.calculate_total_distance()
        
        while current_iteration < max_iterations:
            improved = False
            
            current_vehicles = self.vehicles.copy()
            
            for i, vehicle1 in enumerate(self.vehicles):
                for j, vehicle2 in enumerate(self.vehicles):
                    if i == j:
                        continue
                        
                    for pos1 in range(1, len(vehicle1.route) - 1):
                        node = vehicle1.route[pos1]
                        
                        for pos2 in range(1, len(vehicle2.route)):
                            temp_v1 = Vehicle(self.vehicle_capacity)
                            temp_v1.route = vehicle1.route[:pos1] + vehicle1.route[pos1+1:]
                            
                            temp_v2 = Vehicle(self.vehicle_capacity)
                            temp_v2.route = vehicle2.route[:pos2] + [node] + vehicle2.route[pos2:]
                            
                            feasible = True
                            for temp_v in [temp_v1, temp_v2]:
                                current_time = 0
                                for k in range(1, len(temp_v.route)):
                                    arrival = current_time + self.distance(temp_v.route[k-1], temp_v.route[k])
                                    if arrival > temp_v.route[k].late:
                                        feasible = False
                                        break
                                    current_time = max(arrival, temp_v.route[k].early) + temp_v.route[k].service_time
                            
                            if feasible:
                                self.vehicles[i] = temp_v1
                                self.vehicles[j] = temp_v2
                                new_distance = self.calculate_total_distance()
                                
                                if new_distance < best_distance:
                                    best_distance = new_distance
                                    improved = True
                                else:
                                    self.vehicles = current_vehicles.copy()
            
            if not improved:
                break
                
            current_iteration += 1
        
        return self.vehicles

    def construct_solution(self):
        best_vehicles = None
        best_distance = float('inf')
        
        for attempt in range(3):
            self.vehicles = []
            self.construct_initial_solution()
            
            self.local_search(max_iterations=100)
            
            current_distance = self.calculate_total_distance()
            if current_distance < best_distance:
                best_distance = current_distance
                best_vehicles = []
                for v in self.vehicles:
                    new_v = Vehicle(v.capacity)
                    new_v.route = v.route.copy()
                    new_v.load = v.load
                    new_v.time = v.time
                    new_v.arrival_times = v.arrival_times.copy()
                    best_vehicles.append(new_v)
        
        self.vehicles = best_vehicles
        return self.vehicles

    def is_route_feasible(self, route):
        time = 0
        load = 0
        
        for i in range(len(route) - 1):
            current_node = route[i]
            next_node = route[i + 1]
            
            arrival_time = time + self.distance(current_node, next_node)
            if arrival_time > next_node.late:
                return False
            
            service_start = max(arrival_time, next_node.early)
            time = service_start + next_node.service_time
            
            load += current_node.demand
            if load > self.vehicle_capacity:
                return False
        
        return time <= self.depot.late

    def update_vehicle_times(self, vehicle):
        time = 0
        vehicle.arrival_times = [0]
        vehicle.load = 0
        
        for i in range(1, len(vehicle.route)):
            prev_node = vehicle.route[i-1]
            curr_node = vehicle.route[i]
            
            arrival_time = time + self.distance(prev_node, curr_node)
            service_start = max(arrival_time, curr_node.early)
            time = service_start + curr_node.service_time
            
            vehicle.arrival_times.append(round(arrival_time, 3))
            vehicle.load += curr_node.demand
        
        vehicle.time = time


    def route_distance(self, route):
        return sum(self.distance(route[i], route[i+1]) for i in range(len(route) - 1))
    
    def randomized_initial_solution(self):
        unassigned = self.nodes[1:]  
        random.shuffle(unassigned)  

        while unassigned:
            vehicle = Vehicle(self.vehicle_capacity)
            vehicle.route = [self.depot]
            vehicle.arrival_times = [0]

            assigned_nodes = []
            for node in unassigned:
                if self.is_feasible(vehicle, node):
                    self.update_vehicle(vehicle, node)
                    assigned_nodes.append(node)

            for node in assigned_nodes:
                unassigned.remove(node)

            if len(vehicle.route) > 1:
                vehicle.route.append(self.depot)
                vehicle.arrival_times.append(
                    round(vehicle.time + self.distance(vehicle.route[-2], self.depot), 3)
                )
                self.vehicles.append(vehicle)

            elif unassigned:
                vehicle = Vehicle(self.vehicle_capacity)
                vehicle.route = [self.depot]
                vehicle.arrival_times = [0]
                node = unassigned[0]
                self.update_vehicle(vehicle, node)
                vehicle.route.append(self.depot)
                vehicle.arrival_times.append(
                    round(vehicle.time + self.distance(node, self.depot), 3)
                )
                self.vehicles.append(vehicle)
                unassigned.remove(node)

    def relocation_best_improvement(self):
        improved = True
        while improved:
            improved = False
            best_improvement = 0
            best_move = None

            for v1_index, vehicle1 in enumerate(self.vehicles):
                for i in range(1, len(vehicle1.route) - 1):  
                    node = vehicle1.route[i]
                    
                    for j in range(1, len(vehicle1.route) - 1):
                        if i != j and i != j - 1:
                            new_route = vehicle1.route[:i] + vehicle1.route[i+1:]
                            new_route = new_route[:j] + [node] + new_route[j:]
                            if self.is_route_feasible(new_route):
                                improvement = self.route_distance(vehicle1.route) - self.route_distance(new_route)
                                if improvement > best_improvement:
                                    best_improvement = improvement
                                    best_move = (v1_index, v1_index, i, j)

                    for v2_index, vehicle2 in enumerate(self.vehicles):
                        if v1_index != v2_index:
                            for j in range(1, len(vehicle2.route)):
                                new_route1 = vehicle1.route[:i] + vehicle1.route[i+1:]
                                new_route2 = vehicle2.route[:j] + [node] + vehicle2.route[j:]
                                if self.is_route_feasible(new_route1) and self.is_route_feasible(new_route2):
                                    improvement = (self.route_distance(vehicle1.route) + self.route_distance(vehicle2.route)) - \
                                                  (self.route_distance(new_route1) + self.route_distance(new_route2))
                                    if improvement > best_improvement:
                                        best_improvement = improvement
                                        best_move = (v1_index, v2_index, i, j)

            if best_move:
                v1_index, v2_index, i, j = best_move
                node = self.vehicles[v1_index].route[i]
                
                del self.vehicles[v1_index].route[i]
                self.vehicles[v1_index].load -= node.demand
                
                self.vehicles[v2_index].route.insert(j, node)
                self.vehicles[v2_index].load += node.demand
                
                self.update_vehicle_times(self.vehicles[v1_index])
                self.update_vehicle_times(self.vehicles[v2_index])
                
                improved = True

    def get_elapsed_time(self, start_time):
        return time.time() - start_time
        
    def check_time_limit(self, time_limit, start_time):
        return self.get_elapsed_time(start_time) < time_limit
    
    def construct_solution_multistart(self, time_limit):
        best_vehicles = None
        best_distance = float('inf')
        start_time = time.time()

        start = random.randint(0,4)
        while self.check_time_limit(time_limit, start_time):
            if start % 3 == 0:
                self.construct_initial_solution()
            elif start % 3 == 1:
                self.relocation_best_improvement()
            else:
                self.randomized_initial_solution()

            self.local_search(max_iterations=200)

            tabu_search = TabuSearch(self, tabu_tenure=10, max_iterations=100)
            current_distance = tabu_search.run()

            self.relocation_best_improvement()

            current_distance = self.calculate_total_distance()
            print(f"Iteración {start}: Distancia actual = {current_distance}")

            if current_distance < best_distance:
                best_distance = current_distance
                best_vehicles = [Vehicle(vehicle.capacity) for vehicle in self.vehicles]
                for i, v in enumerate(self.vehicles):
                    best_vehicles[i].route = v.route.copy()
                    best_vehicles[i].load = v.load
                    best_vehicles[i].time = v.time
                    best_vehicles[i].arrival_times = v.arrival_times.copy()

        self.vehicles = best_vehicles
        return self.vehicles

def read_instance(filename):
    with open(filename, 'r') as f:
        n, capacity = map(int, f.readline().split())
        nodes = []
        for _ in range(n + 1):
            id, x, y, demand, early, late, service_time = map(float, f.readline().split())
            nodes.append(Node(int(id), x, y, demand, early, late, service_time))
    return nodes, capacity

def save_results_to_excel(instance_name, vehicles, total_distance, computation_time, vehicle_capacity):
    if not os.path.exists(OUTPUT_FILE):
        workbook = Workbook()
        workbook.remove(workbook.active) 
    else:
        workbook = openpyxl.load_workbook(OUTPUT_FILE)

    sheet = workbook.create_sheet(title=instance_name)

    used_vehicles = [vehicle for vehicle in vehicles if len(vehicle.route) > 2]

    sheet.append([len(used_vehicles), round(total_distance, 3), round(computation_time, 3)])
    
    for vehicle in used_vehicles:
        route = [node.id for node in vehicle.route]
        adjusted_times = [max(at, node.early) for at, node in zip(vehicle.arrival_times, vehicle.route)]
        route_info = [len(route) - 2] + route + adjusted_times + [vehicle.load]
        sheet.append(route_info)

    workbook.save(OUTPUT_FILE)

def read_time_limit(instance_name):
    wb = openpyxl.load_workbook('TimeLimit.xlsx')
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == instance_name:
            return float(row[1].value)
    return 60  

def solve_instance(instance_path):
    nodes, capacity = read_instance(instance_path)
    vrptw = VRPTW(nodes, capacity)

    time_limit = read_time_limit(time_instance)
    start_time = time.time()
    solution = vrptw.construct_solution_multistart(time_limit)  
    end_time = time.time()

    computation_time = (end_time - start_time) * 1000  
    total_distance = vrptw.calculate_total_distance()

    routes = [(vehicle.route, vehicle.arrival_times[-1], vehicle.load) for vehicle in solution]

    return solution, total_distance, computation_time, routes, capacity

def main():
    for instance_filename in os.listdir(INSTANCES_DIR):
        if instance_filename.endswith('.txt'):
            instance_name = instance_filename.replace('.txt', '')
            
            
            instance_path = os.path.join(INSTANCES_DIR, instance_filename)
                
            print(f"\nResolviendo instancia: {instance_name}")
            vehicles, total_distance, computation_time, routes, vehicle_capacity = solve_instance(instance_path)
                
            print(f"Vehículos utilizados: {len(vehicles)}")
            print(f"Distancia total: {total_distance}")
            print(f"Tiempo de cómputo: {computation_time:.2f} ms")
            print(f"Capacidad del vehículo: {vehicle_capacity}")
            print("\nRutas:")
            for i, (route, arrival_time, load) in enumerate(routes, 1):
                print(f"  Ruta {i}:")
                print(f"    Nodos: {' -> '.join(str(node.id) for node in route)}")
                print(f"    Tiempo de llegada: {arrival_time:.2f}")
                print(f"    Carga: {load}")
                
            save_results_to_excel(instance_name, vehicles, total_distance, computation_time, vehicle_capacity)


if __name__ == '__main__':
    main()