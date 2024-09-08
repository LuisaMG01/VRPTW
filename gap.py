import openpyxl

# Función para leer los valores de los archivos Excel por hoja
def read_excel_data_by_sheet(filename):
    workbook = openpyxl.load_workbook(filename)
    sheet_data = {}

    for sheet in workbook.sheetnames:
        sheet_data[sheet] = {}
        sheet_obj = workbook[sheet]
        for row in sheet_obj.iter_rows(min_row=2, values_only=True):
            instance_name = row[0]
            total_distance = row[1]
            lower_bound = row[2]

            try:
                total_distance = float(total_distance)  # Convertir a float
                lower_bound = float(lower_bound)         # Convertir a float
            except (ValueError, TypeError):
                total_distance = None
                lower_bound = None

            if instance_name:
                sheet_data[sheet][instance_name] = {
                    'total_distance': total_distance,
                    'lower_bound': lower_bound
                }
    
    return sheet_data

# Función para calcular el GAP
def calculate_gap(total_distance, lower_bound):
    if lower_bound is None or lower_bound == 0:
        return None  # Evitamos división por cero y valores nulos
    return round(((total_distance - lower_bound) / lower_bound) * 100, 2)

# Función principal para leer los datos y calcular el GAP
def calculate_gaps(constructive_solution_file, lower_bound_file, output_file):
    constructive_data = read_excel_data_by_sheet(constructive_solution_file)
    lower_bound_data = read_excel_data_by_sheet(lower_bound_file)
    
    # Asegúrate de que ambas listas de hojas tienen el mismo número de hojas
    if len(constructive_data) != len(lower_bound_data):
        print("Error: Los archivos tienen un número diferente de hojas.")
        return

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Instance Name", "Total Distance", "Lower Bound", "GAP (%)", "Sheet Index"])

    # Iterar sobre cada hoja en los archivos constructivos
    for i, sheet_name in enumerate(constructive_data.keys()):
        if sheet_name in lower_bound_data:
            constructive_sheet = constructive_data[sheet_name]
            lower_bound_sheet = lower_bound_data[sheet_name]

            for instance_name, values in constructive_sheet.items():
                total_distance = values['total_distance']
                lower_bound = lower_bound_sheet.get(instance_name, {}).get('lower_bound', None)
                
                if total_distance is not None and lower_bound is not None:
                    gap = calculate_gap(total_distance, lower_bound)
                    sheet.append([instance_name, total_distance, lower_bound, gap, i + 1])
                else:
                    sheet.append([instance_name, total_distance if total_distance is not None else "N/A", 
                                  lower_bound if lower_bound is not None else "N/A", 
                                  "N/A", i + 1])  # En caso de no encontrar la cota inferior

    workbook.save(output_file)
    print(f"GAP calculations saved to {output_file}")

# Ejecutar la función con los archivos correspondientes
constructive_solution_file = 'VRPTW_LuisaMariaAlvarez_constructive.xlsx'
lower_bound_file = 'VRPTW_LuisaMariaAlvarez_lowerBound.xlsx'
output_file = 'GAP_results.xlsx'

calculate_gaps(constructive_solution_file, lower_bound_file, output_file)
