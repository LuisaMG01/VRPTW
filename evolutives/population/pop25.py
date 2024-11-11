import random
import copy
from typing import List, Tuple
import os
import openpyxl
from openpyxl import Workbook
import math
import time

OUTPUT_FILE = 'pop25.xlsx'
INSTANCES_DIR = 'instances'  

class EvolutionaryLocalSearch:
    def __init__(self, vrptw, iterations=100, population_size=25, local_search_iterations=20):
        self.vrptw = vrptw
        self.iterations = iterations
        self.population_size = population_size
        self.local_search_iterations = local_search_iterations
        self.best_solution = None
        self.best_fitness = float('inf')

    def local_search(self, solution):
        """Apply local search operators to improve a solution"""
        best_local = copy.deepcopy(solution)
        best_local_fitness = self.calculate_fitness(best_local)

        for _ in range(self.local_search_iterations):
            current = copy.deepcopy(best_local)
            
            # Obtener rutas válidas (con más de 2 nodos)
            valid_routes = [v for v in current if len(v.route) > 2]
            
            if len(valid_routes) >= 2:
                operators = ['relocate', '2-opt', 'exchange']
            elif len(valid_routes) == 1:
                operators = ['2-opt']  # Solo 2-opt si hay una ruta
            else:
                break  # No hay rutas válidas para mejorar
                
            operator = random.choice(operators)

            if operator == 'relocate':
                self._relocate_move(current)
            elif operator == '2-opt':
                self._2opt_move(current)
            else:
                self._exchange_move(current)

            current_fitness = self.calculate_fitness(current)
            if current_fitness < best_local_fitness and self._is_feasible(current):
                best_local = current
                best_local_fitness = current_fitness

        return best_local

    def _relocate_move(self, solution):
        """Move a customer from one route to another"""
        valid_routes = [v for v in solution if len(v.route) > 2]
        if len(valid_routes) < 2:
            return

        # Seleccionar dos rutas aleatorias diferentes
        try:
            route1, route2 = random.sample(valid_routes, 2)
        except ValueError:
            return
        
        if len(route1.route) <= 3:  # Necesitamos al menos un cliente para mover
            return

        # Seleccionar un cliente aleatorio de route1 (excluyendo depósito)
        node_idx = random.randint(1, len(route1.route) - 2)
        node = route1.route.pop(node_idx)
        
        # Insertar en la mejor posición factible de route2
        best_pos = 1
        best_cost = float('inf')
        original_route2 = copy.deepcopy(route2.route)
        
        for i in range(1, len(route2.route)):
            route2.route.insert(i, node)
            if self._is_route_feasible(route2):
                cost = self._calculate_route_cost(route2)
                if cost < best_cost:
                    best_pos = i
                    best_cost = cost
            route2.route = copy.deepcopy(original_route2)
        
        route2.route.insert(best_pos, node)
        self._update_route_info(route1)
        self._update_route_info(route2)

    def _2opt_move(self, solution):
        """Apply 2-opt move to improve a single route"""
        valid_routes = [v for v in solution if len(v.route) > 3]
        if not valid_routes:
            return

        # Seleccionar una ruta aleatoria
        vehicle = random.choice(valid_routes)
        original_route = copy.deepcopy(vehicle.route)
        best_distance = self._calculate_route_cost(vehicle)
        best_route = None
        
        # Intentar diferentes combinaciones de 2-opt
        for i in range(1, len(original_route) - 2):
            for j in range(i + 1, len(original_route) - 1):
                # Revertir el segmento entre i y j
                new_route = original_route[:i] + list(reversed(original_route[i:j+1])) + original_route[j+1:]
                vehicle.route = new_route
                
                if self._is_route_feasible(vehicle):
                    current_distance = self._calculate_route_cost(vehicle)
                    if current_distance < best_distance:
                        best_distance = current_distance
                        best_route = copy.deepcopy(new_route)
                
                vehicle.route = copy.deepcopy(original_route)
        
        if best_route:
            vehicle.route = best_route
            self._update_route_info(vehicle)

    def _exchange_move(self, solution):
        """Exchange customers between two routes"""
        valid_routes = [v for v in solution if len(v.route) > 2]
        if len(valid_routes) < 2:
            return

        try:
            route1, route2 = random.sample(valid_routes, 2)
        except ValueError:
            return

        # Guardar rutas originales
        original_route1 = copy.deepcopy(route1.route)
        original_route2 = copy.deepcopy(route2.route)
        
        best_cost = self._calculate_route_cost(route1) + self._calculate_route_cost(route2)
        best_exchange = None
        
        # Probar todos los posibles intercambios
        for i in range(1, len(route1.route) - 1):
            for j in range(1, len(route2.route) - 1):
                # Intercambiar nodos
                route1.route[i], route2.route[j] = route2.route[j], route1.route[i]
                
                if self._is_route_feasible(route1) and self._is_route_feasible(route2):
                    current_cost = self._calculate_route_cost(route1) + self._calculate_route_cost(route2)
                    if current_cost < best_cost:
                        best_cost = current_cost
                        best_exchange = (i, j)
                
                # Restaurar rutas originales
                route1.route = copy.deepcopy(original_route1)
                route2.route = copy.deepcopy(original_route2)
        
        if best_exchange:
            i, j = best_exchange
            route1.route[i], route2.route[j] = route2.route[j], route1.route[i]
            self._update_route_info(route1)
            self._update_route_info(route2)

    def _is_feasible(self, solution):
        """Check if entire solution is feasible"""
        return all(self._is_route_feasible(vehicle) for vehicle in solution)

    def _is_route_feasible(self, vehicle):
        """Check if a single route is feasible"""
        if not vehicle.route:
            return True

        current_time = 0
        current_load = 0
        
        for i in range(len(vehicle.route) - 1):
            current = vehicle.route[i]
            next_node = vehicle.route[i + 1]
            
            # Actualizar tiempo
            travel_time = self.vrptw.distance(current, next_node)
            arrival_time = current_time + travel_time
            
            # Verificar ventana de tiempo
            if arrival_time > next_node.late:
                return False
            
            service_start = max(arrival_time, next_node.early)
            current_time = service_start + next_node.service_time
            
            # Verificar capacidad
            if i < len(vehicle.route) - 1:  # No considerar demanda del depósito
                current_load += next_node.demand
                if current_load > vehicle.capacity:
                    return False

        return True

    def _update_route_info(self, vehicle):
        """Update route information (times, load)"""
        current_time = 0
        current_load = 0
        arrival_times = [0]
        
        for i in range(len(vehicle.route) - 1):
            current = vehicle.route[i]
            next_node = vehicle.route[i + 1]
            
            travel_time = self.vrptw.distance(current, next_node)
            arrival_time = current_time + travel_time
            service_start = max(arrival_time, next_node.early)
            current_time = service_start + next_node.service_time
            
            arrival_times.append(round(arrival_time, 3))
            if i < len(vehicle.route) - 1:
                current_load += next_node.demand

        vehicle.time = current_time
        vehicle.load = current_load
        vehicle.arrival_times = arrival_times

    def _calculate_route_cost(self, vehicle):
        """Calculate total distance of a route"""
        total_distance = 0
        route = vehicle.route
        for i in range(len(route) - 1):
            total_distance += self.vrptw.distance(route[i], route[i + 1])
        return total_distance

    def calculate_fitness(self, solution):
        """Calculate fitness of a solution"""
        num_vehicles = len([v for v in solution if len(v.route) > 2])
        total_distance = sum(self._calculate_route_cost(v) for v in solution)
        return num_vehicles * 10000 + total_distance

    def optimize(self, initial_solution):
        """Run the ELS optimization"""
        self.best_solution = copy.deepcopy(initial_solution)
        self.best_fitness = self.calculate_fitness(self.best_solution)

        for iteration in range(self.iterations):
            # Generar población inicial
            population = [copy.deepcopy(self.best_solution)]
            
            # Aplicar mutaciones para generar el resto de la población
            while len(population) < self.population_size:
                mutated = copy.deepcopy(self.best_solution)
                for _ in range(random.randint(1, 3)):
                    operator = random.choice(['relocate', '2-opt', 'exchange'])
                    if operator == 'relocate':
                        self._relocate_move(mutated)
                    elif operator == '2-opt':
                        self._2opt_move(mutated)
                    else:
                        self._exchange_move(mutated)
                population.append(mutated)

            # Aplicar búsqueda local a cada solución
            for i in range(len(population)):
                improved = self.local_search(population[i])
                if self._is_feasible(improved):
                    population[i] = improved

            # Actualizar mejor solución
            for solution in population:
                fitness = self.calculate_fitness(solution)
                if fitness < self.best_fitness and self._is_feasible(solution):
                    self.best_solution = copy.deepcopy(solution)
                    self.best_fitness = fitness

            # Imprimir progreso
            if (iteration + 1) % 10 == 0:
                print(f"Iteración ELS {iteration + 1}/{self.iterations}")
                print(f"Mejor fitness: {self.best_fitness}")
                print(f"Número de vehículos: {len([v for v in self.best_solution if len(v.route) > 2])}")
                print(f"Distancia total: {sum(self._calculate_route_cost(v) for v in self.best_solution)}\n")

        return self.best_solution, self.best_fitness

class Node:
    def __init__(self, id, x, y, demand, early, late, service_time):
        self.id = id
        self.x = x
        self.y = y
        self.demand = demand
        self.early = early
        self.late = late
        self.service_time = service_time

class Vehicle:
    def __init__(self, capacity):
        self.capacity = capacity
        self.route = []
        self.load = 0
        self.time = 0
        self.arrival_times = []



class VRPTW:
    def __init__(self, nodes, vehicle_capacity):
        self.nodes = nodes
        self.vehicle_capacity = vehicle_capacity
        self.depot = nodes[0]
        self.vehicles = []

    def print_distance_matrix(self):
        n = len(self.nodes)
        distance_matrix = [[0] * n for _ in range(n)]
        
        for i in range(n):
            for j in range(n):
                distance_matrix[i][j] = round(math.sqrt((self.nodes[i].x - self.nodes[j].x)**2 + (self.nodes[i].y - self.nodes[j].y)**2), 3)
        
        print("Matriz de Distancias:")
        for row in distance_matrix:
            print(row)

    def distance(self, node1, node2):
        return round(math.sqrt((node1.x - node2.x)**2 + (node1.y - node2.y)**2), 3)

    def calculate_total_distance(self):
        total_distance = 0
        for vehicle in self.vehicles:
            route = vehicle.route
            for i in range(len(route) - 1):
                total_distance += self.distance(route[i], route[i + 1])
        return round(total_distance, 3)

    def is_feasible(self, vehicle, node):
        if vehicle.load + node.demand > vehicle.capacity:
            return False
        
        arrival_time = vehicle.time + self.distance(vehicle.route[-1], node)
        
        if arrival_time > node.late:
            return False
        return_time_to_depot = max(arrival_time, node.early) + node.service_time + self.distance(node, self.depot)

        if return_time_to_depot > self.depot.late:
            return False
        
        return True

    def update_vehicle(self, vehicle, node):
        arrival_time = vehicle.time + self.distance(vehicle.route[-1], node)
        service_start = max(arrival_time, node.early)
        vehicle.time = service_start + node.service_time
        
        print(f"Node {node.id} added to vehicle {vehicle}")
        print(f"Arrival Time: {arrival_time}")
        print(f"Service Start: {service_start}")
        print(f"Vehicle Time: {vehicle.time}")
        print(f"Vehicle Load: {vehicle.load}")
        print(f"Vehicle Route: {[node.id for node in vehicle.route]}")
        print(f"Vehicle Arrival Times: {vehicle.arrival_times}")
        print()
        
        vehicle.route.append(node)
        vehicle.load += node.demand
        vehicle.arrival_times.append(round(arrival_time, 3))

    def route_distance(self, route):
        return sum(self.distance(route[i], route[i+1]) for i in range(len(route) - 1))

    def update_vehicle_times(self, vehicle):
        time = 0
        vehicle.arrival_times = [0]
        vehicle.load = 0
        
        for i in range(1, len(vehicle.route)):
            prev_node = vehicle.route[i-1]
            curr_node = vehicle.route[i]
            
            arrival_time = time + self.distance(prev_node, curr_node)
            service_start = max(arrival_time, curr_node.early)
            time = service_start + curr_node.service_time
            
            vehicle.arrival_times.append(round(arrival_time, 3))
            vehicle.load += curr_node.demand
        
        vehicle.time = time

    def relocation_best_improvement(self):
        improved = True
        while improved:
            improved = False
            best_improvement = 0
            best_move = None

            for v1_index, vehicle1 in enumerate(self.vehicles):
                for i in range(1, len(vehicle1.route) - 1):  # Exclude depot
                    node = vehicle1.route[i]
                    
                    # Try relocating within the same route
                    for j in range(1, len(vehicle1.route) - 1):
                        if i != j and i != j - 1:
                            new_route = vehicle1.route[:i] + vehicle1.route[i+1:]
                            new_route = new_route[:j] + [node] + new_route[j:]
                            if self.is_route_feasible(new_route):
                                improvement = self.route_distance(vehicle1.route) - self.route_distance(new_route)
                                if improvement > best_improvement:
                                    best_improvement = improvement
                                    best_move = (v1_index, v1_index, i, j)

                    # Try relocating to other routes
                    for v2_index, vehicle2 in enumerate(self.vehicles):
                        if v1_index != v2_index:
                            for j in range(1, len(vehicle2.route)):
                                new_route1 = vehicle1.route[:i] + vehicle1.route[i+1:]
                                new_route2 = vehicle2.route[:j] + [node] + vehicle2.route[j:]
                                if self.is_route_feasible(new_route1) and self.is_route_feasible(new_route2):
                                    improvement = (self.route_distance(vehicle1.route) + self.route_distance(vehicle2.route)) - \
                                                  (self.route_distance(new_route1) + self.route_distance(new_route2))
                                    if improvement > best_improvement:
                                        best_improvement = improvement
                                        best_move = (v1_index, v2_index, i, j)

            if best_move:
                v1_index, v2_index, i, j = best_move
                node = self.vehicles[v1_index].route[i]
                
                # Remove node from its original position
                del self.vehicles[v1_index].route[i]
                self.vehicles[v1_index].load -= node.demand
                
                # Insert node into its new position
                self.vehicles[v2_index].route.insert(j, node)
                self.vehicles[v2_index].load += node.demand
                
                # Update times for both affected routes
                self.update_vehicle_times(self.vehicles[v1_index])
                self.update_vehicle_times(self.vehicles[v2_index])
                
                improved = True

    def construct_solution(self):
        unassigned = self.nodes[1:]
        max_attempts = len(self.nodes) * 2
        attempts = 0
        
        while unassigned and attempts < max_attempts:
            vehicle = Vehicle(self.vehicle_capacity)
            vehicle.route = [self.depot]
            vehicle.time = 0
            vehicle.arrival_times = [0]
            assigned_any = False
            
            while True:
                feasible_nodes = [node for node in unassigned if self.is_feasible(vehicle, node)]
                
                if not feasible_nodes:
                    break
                
                next_node = min(feasible_nodes, key=lambda node: self.distance(vehicle.route[-1], node))
                
                self.update_vehicle(vehicle, next_node)
                unassigned.remove(next_node)
                assigned_any = True
            
            vehicle.route.append(self.depot)
            vehicle.arrival_times.append(round(vehicle.time + self.distance(vehicle.route[-2], self.depot), 3))
            self.vehicles.append(vehicle)
            
            if not assigned_any:
                attempts += 1
            else:
                attempts = 0
        
        return self.vehicles
    
    def is_route_feasible(self, route):
        time = 0
        load = 0
        
        for i in range(len(route) - 1):
            current_node = route[i]
            next_node = route[i + 1]
            
            arrival_time = time + self.distance(current_node, next_node)
            if arrival_time > next_node.late:
                return False
            
            service_start = max(arrival_time, next_node.early)
            time = service_start + next_node.service_time
            
            load += current_node.demand
            if load > self.vehicle_capacity:
                return False
        
        return time <= self.depot.late
    
class GeneticVRPTW:
    def __init__(self, vrptw, population_size=25, generations=100, mutation_rate=0.1, elite_size=5):
        self.vrptw = vrptw
        self.population_size = population_size
        self.generations = generations
        self.mutation_rate = mutation_rate
        self.elite_size = elite_size
        self.best_solution = None
        self.best_fitness = float('inf')

    def create_initial_population(self, constructive_solution: List[Vehicle]) -> List[List[Vehicle]]:
        """Create initial population using constructive solution and variations"""
        population = []
        
        # Añadir la solución constructiva directamente
        population.append(copy.deepcopy(constructive_solution))
        
        # Generar el resto de la población basada en la solución constructiva
        while len(population) < self.population_size:
            # Crear una variación de la solución constructiva
            new_solution = copy.deepcopy(constructive_solution)
            
            # Aplicar perturbaciones aleatorias
            for _ in range(random.randint(1, 5)):
                new_solution = self.mutate(new_solution)
            
            population.append(new_solution)
        
        return population

    def fitness(self, solution: List[Vehicle]) -> float:
        """Calculate fitness based on number of vehicles and total distance"""
        num_vehicles = len([v for v in solution if len(v.route) > 2])
        total_distance = sum(
            self.vrptw.distance(route[i], route[i + 1])
            for vehicle in solution
            for route in [vehicle.route]
            for i in range(len(route) - 1)
        )
        return num_vehicles * 10000 + total_distance  # Priorizar minimizar vehículos

    def select_parents(self, population: List[List[Vehicle]], fitnesses: List[float]) -> List[List[Vehicle]]:
        """Select parents using tournament selection"""
        tournament_size = 3
        parents = []
        
        for _ in range(self.population_size):
            # Seleccionar índices aleatorios para el torneo
            tournament_indices = random.sample(range(len(population)), tournament_size)
            # Encontrar el índice del ganador basado en fitness
            winner_idx = min(tournament_indices, key=lambda idx: fitnesses[idx])
            parents.append(copy.deepcopy(population[winner_idx]))
        
        return parents

    def crossover(self, parent1: List[Vehicle], parent2: List[Vehicle]) -> List[Vehicle]:
        """Perform route-based crossover between two parents"""
        if not parent1 or not parent2:
            return parent1 if parent1 else parent2

        # Seleccionar rutas aleatorias de cada padre
        routes1 = [v.route[1:-1] for v in parent1 if len(v.route) > 2]
        routes2 = [v.route[1:-1] for v in parent2 if len(v.route) > 2]
        
        if not routes1 or not routes2:
            return parent1 if routes1 else parent2

        # Crear hijo combinando rutas
        child_routes = []
        used_nodes = set()
        
        # Añadir algunas rutas del primer padre
        for route in routes1[:len(routes1)//2]:
            route_nodes = [node.id for node in route]
            if not any(node_id in used_nodes for node_id in route_nodes):
                child_routes.append(route)
                used_nodes.update(route_nodes)

        # Añadir nodos restantes usando rutas del segundo padre
        for route in routes2:
            route_nodes = [node.id for node in route]
            if not any(node_id in used_nodes for node_id in route_nodes):
                child_routes.append(route)
                used_nodes.update(route_nodes)

        # Convertir rutas a vehículos
        child_vehicles = []
        for route in child_routes:
            vehicle = Vehicle(self.vrptw.vehicle_capacity)
            vehicle.route = [self.vrptw.depot] + list(route) + [self.vrptw.depot]
            
            # Recalcular tiempos y cargas
            vehicle.time = 0
            vehicle.load = 0
            vehicle.arrival_times = [0]
            
            for i in range(1, len(vehicle.route)):
                node = vehicle.route[i]
                prev_node = vehicle.route[i-1]
                arrival_time = vehicle.time + self.vrptw.distance(prev_node, node)
                service_start = max(arrival_time, node.early) if i < len(vehicle.route)-1 else arrival_time
                vehicle.time = service_start + (node.service_time if i < len(vehicle.route)-1 else 0)
                vehicle.arrival_times.append(round(arrival_time, 3))
                vehicle.load += node.demand if i < len(vehicle.route)-1 else 0
                
            child_vehicles.append(vehicle)

        return child_vehicles

    def mutate(self, solution: List[Vehicle]) -> List[Vehicle]:
        """Apply mutation operators to the solution"""
        if random.random() < self.mutation_rate:
            mutation_type = random.choice(['swap', 'insert', 'reverse'])
            
            if mutation_type == 'swap':
                # Intercambiar dos nodos aleatorios entre rutas
                routes = [v.route[1:-1] for v in solution if len(v.route) > 2]
                if len(routes) >= 2:
                    route1, route2 = random.sample(routes, 2)
                    if route1 and route2:
                        pos1 = random.randint(0, len(route1)-1)
                        pos2 = random.randint(0, len(route2)-1)
                        route1[pos1], route2[pos2] = route2[pos2], route1[pos1]
            
            elif mutation_type == 'insert':
                # Mover un nodo aleatorio a una posición diferente
                routes = [v.route[1:-1] for v in solution if len(v.route) > 2]
                if routes:
                    route = random.choice(routes)
                    if len(route) > 1:
                        pos1 = random.randint(0, len(route)-1)
                        pos2 = random.randint(0, len(route)-1)
                        node = route.pop(pos1)
                        route.insert(pos2, node)
            
            else:  # reverse
                # Invertir una subsecuencia aleatoria en una ruta
                routes = [v.route[1:-1] for v in solution if len(v.route) > 2]
                if routes:
                    route = random.choice(routes)
                    if len(route) > 2:
                        pos1 = random.randint(0, len(route)-2)
                        pos2 = random.randint(pos1+1, len(route)-1)
                        route[pos1:pos2+1] = reversed(route[pos1:pos2+1])

        return solution

    def optimize(self, constructive_solution: List[Vehicle]) -> Tuple[List[Vehicle], float]:
        """Run the genetic algorithm optimization"""
        # Crear población inicial usando la solución constructiva
        population = self.create_initial_population(constructive_solution)
        
        # Bucle principal del GA
        for generation in range(self.generations):
            # Calcular fitness para todas las soluciones
            fitnesses = [self.fitness(solution) for solution in population]
            
            # Actualizar mejor solución
            min_fitness_idx = min(range(len(fitnesses)), key=fitnesses.__getitem__)
            if fitnesses[min_fitness_idx] < self.best_fitness:
                self.best_fitness = fitnesses[min_fitness_idx]
                self.best_solution = copy.deepcopy(population[min_fitness_idx])
            
            # Seleccionar padres
            parents = self.select_parents(population, fitnesses)
            
            # Crear nueva población
            new_population = []
            
            # Elitismo: mantener mejores soluciones
            # Crear lista de índices ordenados por fitness
            sorted_indices = sorted(range(len(fitnesses)), key=lambda k: fitnesses[k])
            elite_solutions = [population[i] for i in sorted_indices[:self.elite_size]]
            new_population.extend(copy.deepcopy(elite_solutions))
            
            # Crear descendencia
            while len(new_population) < self.population_size:
                parent1, parent2 = random.sample(parents, 2)
                child = self.crossover(parent1, parent2)
                child = self.mutate(child)
                new_population.append(child)
            
            population = new_population
            
            # Imprimir progreso
            if (generation + 1) % 10 == 0:
                print(f"Generación {generation + 1}/{self.generations}")
                print(f"Mejor fitness: {self.best_fitness}")
                print(f"Número de vehículos: {len([v for v in self.best_solution if len(v.route) > 2])}")
                print(f"Distancia total: {sum(self.vrptw.distance(route[i], route[i + 1]) for vehicle in self.best_solution for route in [vehicle.route] for i in range(len(route) - 1))}")
        
        return self.best_solution, self.best_fitness

def solve_instance_with_ga(instance_path):
    """Solve VRPTW instance using constructive solution, genetic algorithm, and ELS"""
    # Obtener la solución constructiva
    nodes, capacity = read_instance(instance_path)
    vrptw = VRPTW(nodes, capacity)
    
    print("Generando solución constructiva inicial...")
    constructive_solution = vrptw.construct_solution()
    
    # Aplicar mejora por reubicación
    print("Aplicando mejora por reubicación...")
    vrptw.relocation_best_improvement()  # Modifica la solución in-place
    constructive_distance = vrptw.calculate_total_distance()
    
    # Usar los vehículos del VRPTW como solución constructiva
    constructive_solution = vrptw.vehicles
    
    print(f"Solución constructiva: {len(constructive_solution)} vehículos, distancia: {constructive_distance}")
    
    # Aplicar el algoritmo genético
    print("\nAplicando algoritmo genético...")
    ga = GeneticVRPTW(vrptw)
    start_time = time.time()
    ga_solution, ga_fitness = ga.optimize(constructive_solution)
    
    # Aplicar ELS para mejorar la solución
    print("\nAplicando búsqueda local evolutiva...")
    els = EvolutionaryLocalSearch(vrptw)
    best_solution, best_fitness = els.optimize(ga_solution)
    
    computation_time = (time.time() - start_time) * 1000
    
    total_distance = sum(
        vrptw.distance(route[i], route[i + 1])
        for vehicle in best_solution
        for route in [vehicle.route]
        for i in range(len(route) - 1)
    )
    
    routes = [(vehicle.route, vehicle.arrival_times[-1], vehicle.load) 
              for vehicle in best_solution]
    
    return best_solution, total_distance, computation_time, routes, capacity

def update_vehicle_times(self, vehicle):
    """Update the arrival times and total time for a vehicle's route"""
    time = 0
    arrival_times = [0]  # Tiempo de llegada al depósito inicial
    
    for i in range(1, len(vehicle.route)):
        current_node = vehicle.route[i-1]
        next_node = vehicle.route[i]
        
        arrival_time = time + self.distance(current_node, next_node)
        service_start = max(arrival_time, next_node.early) if i < len(vehicle.route)-1 else arrival_time
        time = service_start + (next_node.service_time if i < len(vehicle.route)-1 else 0)
        
        arrival_times.append(round(arrival_time, 3))
    
    vehicle.time = time
    vehicle.arrival_times = arrival_times

def read_instance(filename):
    with open(filename, 'r') as f:
        n, capacity = map(int, f.readline().split())
        nodes = []
        for _ in range(n + 1):
            id, x, y, demand, early, late, service_time = map(float, f.readline().split())
            nodes.append(Node(int(id), x, y, demand, early, late, service_time))
    return nodes, capacity

def save_results_to_excel(instance_name, vehicles, total_distance, computation_time, vehicle_capacity):
    if not os.path.exists(OUTPUT_FILE):
        workbook = Workbook()
        workbook.remove(workbook.active) 
    else:
        workbook = openpyxl.load_workbook(OUTPUT_FILE)

    sheet = workbook.create_sheet(title=instance_name)

    used_vehicles = [vehicle for vehicle in vehicles if len(vehicle.route) > 2]

    sheet.append([len(used_vehicles), round(total_distance, 3), round(computation_time, 3)])
    
    for vehicle in used_vehicles:
        route = [node.id for node in vehicle.route]
        adjusted_times = [max(at, node.early) for at, node in zip(vehicle.arrival_times, vehicle.route)]

        route_info = [len(route) - 2] + route + adjusted_times + [vehicle.load]
        sheet.append(route_info)

    workbook.save(OUTPUT_FILE)

def main():
    for instance_filename in os.listdir(INSTANCES_DIR):
        if instance_filename.endswith('.txt'):
            instance_path = os.path.join(INSTANCES_DIR, instance_filename)
            instance_name = instance_filename.replace('.txt', '')
            
            print(f"\nResolviendo instancia: {instance_name}")
            
            vehicles, total_distance, computation_time, routes, vehicle_capacity = (
                solve_instance_with_ga(instance_path)
            )
            
            print(f"\nResultados finales:")
            print(f"Vehículos: {len(vehicles)}")
            print(f"Distancia total: {total_distance}")
            print(f"Tiempo de cómputo: {computation_time:.2f} ms")
            print(f"Capacidad del vehículo: {vehicle_capacity}")
            print("Rutas:")
            for i, (route, arrival_time, load) in enumerate(routes, 1):
                print(f"  Ruta {i}:")
                print(f"    Nodos: {' -> '.join(str(node.id) for node in route)}")
                print(f"    Tiempo de llegada: {arrival_time:.2f}")
                print(f"    Carga: {load}")
            
            save_results_to_excel(instance_name, vehicles, total_distance, 
                                computation_time, vehicle_capacity)

if __name__ == '__main__':
    main()