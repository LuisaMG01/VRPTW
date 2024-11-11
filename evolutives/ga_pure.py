import random
import copy
from typing import List, Tuple
import time
import os
import math
import openpyxl
from openpyxl import Workbook

OUTPUT_FILE = 'pure.xlsx'
INSTANCES_DIR = 'instances'  


class VRPTW:
    def __init__(self, nodes, vehicle_capacity):
        self.nodes = nodes
        self.vehicle_capacity = vehicle_capacity
        self.depot = nodes[0]
        self.vehicles = []

    def print_distance_matrix(self):
        n = len(self.nodes)
        distance_matrix = [[0] * n for _ in range(n)]
        
        for i in range(n):
            for j in range(n):
                distance_matrix[i][j] = round(math.sqrt((self.nodes[i].x - self.nodes[j].x)**2 + (self.nodes[i].y - self.nodes[j].y)**2), 3)
        
        print("Matriz de Distancias:")
        for row in distance_matrix:
            print(row)

    def distance(self, node1, node2):
        return round(math.sqrt((node1.x - node2.x)**2 + (node1.y - node2.y)**2), 3)

    def calculate_total_distance(self):
        total_distance = 0
        for vehicle in self.vehicles:
            route = vehicle.route
            for i in range(len(route) - 1):
                total_distance += self.distance(route[i], route[i + 1])
        return round(total_distance, 3)

    def is_feasible(self, vehicle, node):
        if vehicle.load + node.demand > vehicle.capacity:
            return False
        
        arrival_time = vehicle.time + self.distance(vehicle.route[-1], node)
        
        if arrival_time > node.late:
            return False
        return_time_to_depot = max(arrival_time, node.early) + node.service_time + self.distance(node, self.depot)

        if return_time_to_depot > self.depot.late:
            return False
        
        return True

    def update_vehicle(self, vehicle, node):
        arrival_time = vehicle.time + self.distance(vehicle.route[-1], node)
        service_start = max(arrival_time, node.early)
        vehicle.time = service_start + node.service_time
        
        print(f"Node {node.id} added to vehicle {vehicle}")
        print(f"Arrival Time: {arrival_time}")
        print(f"Service Start: {service_start}")
        print(f"Vehicle Time: {vehicle.time}")
        print(f"Vehicle Load: {vehicle.load}")
        print(f"Vehicle Route: {[node.id for node in vehicle.route]}")
        print(f"Vehicle Arrival Times: {vehicle.arrival_times}")
        print()
        
        vehicle.route.append(node)
        vehicle.load += node.demand
        vehicle.arrival_times.append(round(arrival_time, 3))

    def construct_solution(self):
        unassigned = self.nodes[1:]
        max_attempts = len(self.nodes) * 2
        attempts = 0
        
        while unassigned and attempts < max_attempts:
            vehicle = Vehicle(self.vehicle_capacity)
            vehicle.route = [self.depot]
            vehicle.time = 0
            vehicle.arrival_times = [0]
            assigned_any = False
            
            while True:
                feasible_nodes = [node for node in unassigned if self.is_feasible(vehicle, node)]
                
                if not feasible_nodes:
                    break
                
                next_node = min(feasible_nodes, key=lambda node: self.distance(vehicle.route[-1], node))
                
                self.update_vehicle(vehicle, next_node)
                unassigned.remove(next_node)
                assigned_any = True
            
            vehicle.route.append(self.depot)
            vehicle.arrival_times.append(round(vehicle.time + self.distance(vehicle.route[-2], self.depot), 3))
            self.vehicles.append(vehicle)
            
            if not assigned_any:
                attempts += 1
            else:
                attempts = 0
        
        return self.vehicles
    
class Node:
    def __init__(self, id, x, y, demand, early, late, service_time):
        self.id = id
        self.x = x
        self.y = y
        self.demand = demand
        self.early = early
        self.late = late
        self.service_time = service_time

class Vehicle:
    def __init__(self, capacity):
        self.capacity = capacity
        self.route = []
        self.load = 0
        self.time = 0
        self.arrival_times = []

class GeneticVRPTW(VRPTW):
    def __init__(self, nodes, vehicle_capacity):
        super().__init__(nodes, vehicle_capacity)
        # Parámetros GA optimizados para mejor convergencia
        self.population_size = 150
        self.generations = 400
        self.mutation_rate = 0.3
        self.elite_size = 15
        self.tournament_size = 5
        self.crossover_rate = 0.85

    def generate_random_solution(self) -> List[Vehicle]:
        """Genera una solución aleatoria pero considerando cercanía espacial"""
        unassigned = self.nodes[1:]
        solution = []
        
        while unassigned:
            vehicle = Vehicle(self.vehicle_capacity)
            vehicle.route = [self.depot]
            current_node = self.depot
            current_load = 0
            current_time = 0
            
            # Construir ruta basada en vecinos más cercanos con aleatorización
            while unassigned:
                # Encontrar nodos factibles
                feasible_nodes = []
                for node in unassigned:
                    if current_load + node.demand <= self.vehicle_capacity:
                        arrival_time = current_time + self.distance(current_node, node)
                        if arrival_time <= node.late:
                            feasible_nodes.append(node)
                
                if not feasible_nodes:
                    break
                
                # Seleccionar aleatoriamente entre los k mejores vecinos
                k = min(3, len(feasible_nodes))
                feasible_nodes.sort(key=lambda x: self.distance(current_node, x))
                next_node = random.choice(feasible_nodes[:k])
                
                vehicle.route.append(next_node)
                current_load += next_node.demand
                arrival_time = current_time + self.distance(current_node, next_node)
                current_time = max(arrival_time, next_node.early) + next_node.service_time
                current_node = next_node
                unassigned.remove(next_node)
            
            vehicle.route.append(self.depot)
            solution.append(vehicle)
            
        return solution

    def create_initial_population(self) -> List[List[Vehicle]]:
        """Crea población inicial con mayor diversidad"""
        population = []
        for i in range(self.population_size):
            # 70% constructivo con aleatorización, 30% aleatorio
            if random.random() < 0.7:
                solution = self.construct_randomized_solution()
            else:
                solution = self.generate_random_solution()
            population.append(copy.deepcopy(solution))
        return population
    
    def construct_randomized_solution(self) -> List[Vehicle]:
        """Constructivo con elemento aleatorio controlado"""
        unassigned = self.nodes[1:]
        solution = []
        
        while unassigned:
            vehicle = Vehicle(self.vehicle_capacity)
            vehicle.route = [self.depot]
            current_time = 0
            current_load = 0
            
            while unassigned:
                feasible_nodes = [
                    node for node in unassigned 
                    if current_load + node.demand <= self.vehicle_capacity
                ]
                
                if not feasible_nodes:
                    break
                    
                # Seleccionar entre los k mejores candidatos
                feasible_nodes.sort(key=lambda x: self.distance(vehicle.route[-1], x))
                k = min(5, len(feasible_nodes))
                selected_nodes = feasible_nodes[:k]
                next_node = random.choice(selected_nodes)
                
                vehicle.route.append(next_node)
                current_load += next_node.demand
                unassigned.remove(next_node)
            
            if len(vehicle.route) > 1:
                vehicle.route.append(self.depot)
                solution.append(vehicle)
                
        return solution

    def fitness(self, solution: List[Vehicle]) -> Tuple[float, float]:
        """Función de fitness mejorada con penalizaciones calibradas"""
        total_distance = 0
        penalty = 0
        vehicles_penalty = len(solution) * 500  # Penalización por número de vehículos
        
        for vehicle in solution:
            route = vehicle.route
            current_time = 0
            current_load = 0
            
            for i in range(len(route) - 1):
                current_node = route[i]
                next_node = route[i + 1]
                distance = self.distance(current_node, next_node)
                total_distance += distance
                current_time += distance
                
                if i > 0:  # No aplicar ventana temporal al depósito inicial
                    # Penalización por ventana temporal
                    if current_time < current_node.early:
                        waiting = current_node.early - current_time
                        penalty += waiting * 10
                        current_time = current_node.early
                    elif current_time > current_node.late:
                        delay = current_time - current_node.late
                        penalty += delay * 100
                    
                    current_load += current_node.demand
                
                current_time += current_node.service_time
            
            # Penalización por capacidad
            if current_load > vehicle.capacity:
                penalty += (current_load - vehicle.capacity) * 200
        
        return total_distance + penalty + vehicles_penalty, total_distance

    def tournament_selection(self, population: List[List[Vehicle]]) -> List[Vehicle]:
        """Selección por torneo mejorada con ranking"""
        tournament = random.sample(population, self.tournament_size)
        tournament.sort(key=lambda x: self.fitness(x)[0])
        
        # Probabilidad de selección basada en ranking
        rank_weights = [1/(i+1) for i in range(len(tournament))]
        total_weight = sum(rank_weights)
        probabilities = [w/total_weight for w in rank_weights]
        
        return random.choices(tournament, weights=probabilities, k=1)[0]

    def order_crossover(self, parent1: List[Vehicle], parent2: List[Vehicle]) -> List[Vehicle]:
        """Order Crossover mejorado para VRPTW"""
        # Convertir rutas a una secuencia única
        sequence1 = [node for vehicle in parent1 for node in vehicle.route[1:-1]]
        sequence2 = [node for vehicle in parent2 for node in vehicle.route[1:-1]]
        
        if not sequence1 or not sequence2:
            return copy.deepcopy(parent1)
        
        # Puntos de cruce
        length = min(len(sequence1), len(sequence2))
        if length < 2:
            return copy.deepcopy(parent1)
            
        point1, point2 = sorted(random.sample(range(length), 2))
        
        # Crear hijo preservando el orden
        child_sequence = [None] * length
        # Copiar segmento del primer padre
        segment = sequence1[point1:point2]
        child_sequence[point1:point2] = segment
        
        # Completar con elementos del segundo padre
        remaining = [node for node in sequence2 if node not in segment]
        j = 0
        for i in range(length):
            if child_sequence[i] is None:
                if j < len(remaining):
                    child_sequence[i] = remaining[j]
                    j += 1
                    
        # Construir solución viable
        return self.build_solution_from_sequence(child_sequence)

    def mutation(self, solution: List[Vehicle]) -> List[Vehicle]:
        """Mutación mejorada con múltiples operadores"""
        if random.random() > self.mutation_rate:
            return solution
            
        mutated = copy.deepcopy(solution)
        
        # Aplicar uno de tres operadores de mutación
        mutation_type = random.random()
        
        if mutation_type < 0.33:
            # Swap mutation
            all_nodes = [node for vehicle in mutated for node in vehicle.route[1:-1]]
            if len(all_nodes) >= 2:
                idx1, idx2 = random.sample(range(len(all_nodes)), 2)
                all_nodes[idx1], all_nodes[idx2] = all_nodes[idx2], all_nodes[idx1]
                return self.build_solution_from_sequence(all_nodes)
                
        elif mutation_type < 0.66:
            # Reverse mutation
            if len(mutated) > 0:
                vehicle = random.choice(mutated)
                if len(vehicle.route) > 3:
                    start = random.randint(1, len(vehicle.route)-3)
                    end = random.randint(start+1, len(vehicle.route)-2)
                    vehicle.route[start:end+1] = reversed(vehicle.route[start:end+1])
                    
        else:
            # Relocation mutation
            if len(mutated) > 1:
                source = random.randint(0, len(mutated)-1)
                target = random.randint(0, len(mutated)-1)
                if len(mutated[source].route) > 3:
                    node_idx = random.randint(1, len(mutated[source].route)-2)
                    node = mutated[source].route.pop(node_idx)
                    insert_idx = random.randint(1, len(mutated[target].route)-1)
                    mutated[target].route.insert(insert_idx, node)
        
        return mutated

    def build_solution_from_sequence(self, sequence: List[Node]) -> List[Vehicle]:
        """Construye una solución viable a partir de una secuencia de nodos"""
        solution = []
        current_vehicle = Vehicle(self.vehicle_capacity)
        current_vehicle.route = [self.depot]
        current_load = 0
        
        for node in sequence:
            if current_load + node.demand <= self.vehicle_capacity:
                current_vehicle.route.append(node)
                current_load += node.demand
            else:
                current_vehicle.route.append(self.depot)
                solution.append(current_vehicle)
                current_vehicle = Vehicle(self.vehicle_capacity)
                current_vehicle.route = [self.depot, node]
                current_load = node.demand
        
        if len(current_vehicle.route) > 1:
            current_vehicle.route.append(self.depot)
            solution.append(current_vehicle)
            
        return solution

    def genetic_algorithm(self) -> List[Vehicle]:
        """Algoritmo genético principal optimizado"""
        population = self.create_initial_population()
        best_solution = None
        best_fitness = float('inf')
        generations_without_improvement = 0
        
        for generation in range(self.generations):
            # Evaluar población
            population_fitness = [(solution, self.fitness(solution)) for solution in population]
            population_fitness.sort(key=lambda x: x[1][0])
            
            # Actualizar mejor solución
            current_best = population_fitness[0]
            if current_best[1][0] < best_fitness:
                best_fitness = current_best[1][0]
                best_solution = copy.deepcopy(current_best[0])
                generations_without_improvement = 0
                print(f"Generation {generation}: New best fitness = {best_fitness}")
            else:
                generations_without_improvement += 1
            
            # Criterio de reinicio
            if generations_without_improvement > 50:
                print(f"Generation {generation}: Restarting population")
                new_population = self.create_initial_population()
                # Mantener el mejor
                new_population[0] = copy.deepcopy(best_solution)
                population = new_population
                generations_without_improvement = 0
                continue
            
            # Crear nueva población
            new_population = []
            
            # Elitismo
            new_population.extend(copy.deepcopy(solution) for solution, _ 
                                in population_fitness[:self.elite_size])
            
            # Generar resto de la población
            while len(new_population) < self.population_size:
                if random.random() < self.crossover_rate:
                    parent1 = self.tournament_selection(population)
                    parent2 = self.tournament_selection(population)
                    child = self.order_crossover(parent1, parent2)
                else:
                    child = self.tournament_selection(population)
                
                child = self.mutation(child)
                new_population.append(child)
            
            population = new_population
        
        return best_solution

def solve_instance(instance_path):
    nodes, capacity = read_instance(instance_path)
    vrptw = GeneticVRPTW(nodes, capacity)
    
    start_time = time.time()
    solution = vrptw.genetic_algorithm()
    end_time = time.time()
    
    computation_time = (end_time - start_time) * 1000
    total_distance = vrptw.calculate_total_distance()
    
    routes = [(vehicle.route, vehicle.arrival_times[-1] if vehicle.arrival_times else 0, vehicle.load) 
              for vehicle in solution]
    
    return solution, total_distance, computation_time, routes, capacity

   

def read_instance(filename):
    with open(filename, 'r') as f:
        n, capacity = map(int, f.readline().split())
        nodes = []
        for _ in range(n + 1):
            id, x, y, demand, early, late, service_time = map(float, f.readline().split())
            nodes.append(Node(int(id), x, y, demand, early, late, service_time))
    return nodes, capacity

def save_results_to_excel(instance_name, vehicles, total_distance, computation_time, vehicle_capacity):
    if not os.path.exists(OUTPUT_FILE):
        workbook = Workbook()
        workbook.remove(workbook.active) 
    else:
        workbook = openpyxl.load_workbook(OUTPUT_FILE)

    sheet = workbook.create_sheet(title=instance_name)

    used_vehicles = [vehicle for vehicle in vehicles if len(vehicle.route) > 2]

    sheet.append([len(used_vehicles), round(total_distance, 3), round(computation_time, 3)])
    
    for vehicle in used_vehicles:
        route = [node.id for node in vehicle.route]
        adjusted_times = [max(at, node.early) for at, node in zip(vehicle.arrival_times, vehicle.route)]

        route_info = [len(route) - 2] + route + adjusted_times + [vehicle.load]
        sheet.append(route_info)

    workbook.save(OUTPUT_FILE)



def main():
    for instance_filename in os.listdir(INSTANCES_DIR):
        if instance_filename.endswith('.txt'):
            instance_path = os.path.join(INSTANCES_DIR, instance_filename)
            instance_name = instance_filename.replace('.txt', '')
            vehicles, total_distance, computation_time, routes, vehicle_capacity = solve_instance(instance_path)
            print(f"Vehicles: {len(vehicles)}")
            print(f"Total Distance: {total_distance}")
            print(f"Computation Time: {computation_time:.2f} ms")
            print(f"Vehicle Capacity: {vehicle_capacity}")
            print("Routes:")
            for i, (route, arrival_time, load) in enumerate(routes, 1):
                print(f"  Route {i}:")
                print(f"    Nodes: {' -> '.join(str(node.id) for node in route)}")
                print(f"    Arrival Time: {arrival_time:.2f}")
                print(f"    Load: {load}")
            save_results_to_excel(instance_name, vehicles, total_distance, computation_time, vehicle_capacity)

if __name__ == '__main__':
    main()
