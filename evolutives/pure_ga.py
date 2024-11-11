import random
import copy
from typing import List, Tuple
import os
import openpyxl
from openpyxl import Workbook
import math
import time

OUTPUT_FILE = 'pure_1.xlsx'
INSTANCES_DIR = 'instances'  

class Node:
    def __init__(self, id, x, y, demand, early, late, service_time):
        self.id = id
        self.x = x
        self.y = y
        self.demand = demand
        self.early = early
        self.late = late
        self.service_time = service_time

class Vehicle:
    def __init__(self, capacity):
        self.capacity = capacity
        self.route = []
        self.load = 0
        self.time = 0
        self.arrival_times = []



class VRPTW:
    def __init__(self, nodes, vehicle_capacity):
        self.nodes = nodes
        self.vehicle_capacity = vehicle_capacity
        self.depot = nodes[0]
        self.vehicles = []
        self.RCL_SIZE = 3 

    def print_distance_matrix(self):
        n = len(self.nodes)
        distance_matrix = [[0] * n for _ in range(n)]
        
        for i in range(n):
            for j in range(n):
                distance_matrix[i][j] = round(math.sqrt((self.nodes[i].x - self.nodes[j].x)**2 + (self.nodes[i].y - self.nodes[j].y)**2), 3)
        
        print("Matriz de Distancias:")
        for row in distance_matrix:
            print(row)

    def distance(self, node1, node2):
        return round(math.sqrt((node1.x - node2.x)**2 + (node1.y - node2.y)**2), 3)

    def calculate_total_distance(self):
        total_distance = 0
        for vehicle in self.vehicles:
            route = vehicle.route
            for i in range(len(route) - 1):
                total_distance += self.distance(route[i], route[i + 1])
        return round(total_distance, 3)

    def is_feasible(self, vehicle, node):
        if vehicle.load + node.demand > vehicle.capacity:
            return False
        
        arrival_time = vehicle.time + self.distance(vehicle.route[-1], node)
        
        if arrival_time > node.late:
            return False
        return_time_to_depot = max(arrival_time, node.early) + node.service_time + self.distance(node, self.depot)

        if return_time_to_depot > self.depot.late:
            return False
        
        return True

    def update_vehicle(self, vehicle, node):
        arrival_time = vehicle.time + self.distance(vehicle.route[-1], node)
        service_start = max(arrival_time, node.early)
        vehicle.time = service_start + node.service_time
        
        print(f"Node {node.id} added to vehicle {vehicle}")
        print(f"Arrival Time: {arrival_time}")
        print(f"Service Start: {service_start}")
        print(f"Vehicle Time: {vehicle.time}")
        print(f"Vehicle Load: {vehicle.load}")
        print(f"Vehicle Route: {[node.id for node in vehicle.route]}")
        print(f"Vehicle Arrival Times: {vehicle.arrival_times}")
        print()
        
        vehicle.route.append(node)
        vehicle.load += node.demand
        vehicle.arrival_times.append(round(arrival_time, 3))

    def route_distance(self, route):
        return sum(self.distance(route[i], route[i+1]) for i in range(len(route) - 1))

    def update_vehicle_times(self, vehicle):
        time = 0
        vehicle.arrival_times = [0]
        vehicle.load = 0
        
        for i in range(1, len(vehicle.route)):
            prev_node = vehicle.route[i-1]
            curr_node = vehicle.route[i]
            
            arrival_time = time + self.distance(prev_node, curr_node)
            service_start = max(arrival_time, curr_node.early)
            time = service_start + curr_node.service_time
            
            vehicle.arrival_times.append(round(arrival_time, 3))
            vehicle.load += curr_node.demand
        
        vehicle.time = time

    def construct_solution(self):
        unassigned = self.nodes[1:]
        max_attempts = len(self.nodes) * 2
        attempts = 0
        
        while unassigned and attempts < max_attempts:
            vehicle = Vehicle(self.vehicle_capacity)
            vehicle.route = [self.depot]
            vehicle.time = 0
            vehicle.arrival_times = [0]
            assigned_any = False
            
            while True:
                feasible_nodes = [node for node in unassigned if self.is_feasible(vehicle, node)]
                
                if not feasible_nodes:
                    break
                
               
                feasible_nodes.sort(key=lambda node: self.distance(vehicle.route[-1], node))
                
                rcl_size = min(self.RCL_SIZE, len(feasible_nodes))
                rcl = feasible_nodes[:rcl_size]
                
                next_node = random.choice(rcl)
                
                self.update_vehicle(vehicle, next_node)
                unassigned.remove(next_node)
                assigned_any = True
            
            vehicle.route.append(self.depot)
            vehicle.arrival_times.append(round(vehicle.time + self.distance(vehicle.route[-2], self.depot), 3))
            self.vehicles.append(vehicle)
            
            if not assigned_any:
                attempts += 1
            else:
                attempts = 0
        
        return self.vehicles
    
    def is_route_feasible(self, route):
        time = 0
        load = 0
        
        for i in range(len(route) - 1):
            current_node = route[i]
            next_node = route[i + 1]
            
            arrival_time = time + self.distance(current_node, next_node)
            if arrival_time > next_node.late:
                return False
            
            service_start = max(arrival_time, next_node.early)
            time = service_start + next_node.service_time
            
            load += current_node.demand
            if load > self.vehicle_capacity:
                return False
        
        return time <= self.depot.late
    
class GeneticVRPTW:
    def __init__(self, vrptw, population_size=50, generations=100, mutation_rate=0.1, elite_size=5):
        self.vrptw = vrptw
        self.population_size = population_size
        self.generations = generations
        self.mutation_rate = mutation_rate
        self.elite_size = elite_size
        self.best_solution = None
        self.best_fitness = float('inf')

    def create_initial_population(self, constructive_solution: List[Vehicle]) -> List[List[Vehicle]]:
        population = []
        
        population.append(copy.deepcopy(constructive_solution))
        
        while len(population) < self.population_size:
            new_solution = copy.deepcopy(constructive_solution)
            
            for _ in range(random.randint(1, 5)):
                new_solution = self.mutate(new_solution)
            
            population.append(new_solution)
        
        return population

    def fitness(self, solution: List[Vehicle]) -> float:
        num_vehicles = len([v for v in solution if len(v.route) > 2])
        total_distance = sum(
            self.vrptw.distance(route[i], route[i + 1])
            for vehicle in solution
            for route in [vehicle.route]
            for i in range(len(route) - 1)
        )
        return num_vehicles * 10000 + total_distance  

    def select_parents(self, population: List[List[Vehicle]], fitnesses: List[float]) -> List[List[Vehicle]]:
        tournament_size = 3
        parents = []
        
        for _ in range(self.population_size):
            tournament_indices = random.sample(range(len(population)), tournament_size)
            winner_idx = min(tournament_indices, key=lambda idx: fitnesses[idx])
            parents.append(copy.deepcopy(population[winner_idx]))
        
        return parents

    def crossover(self, parent1: List[Vehicle], parent2: List[Vehicle]) -> List[Vehicle]:
        if not parent1 or not parent2:
            return parent1 if parent1 else parent2

        routes1 = [v.route[1:-1] for v in parent1 if len(v.route) > 2]
        routes2 = [v.route[1:-1] for v in parent2 if len(v.route) > 2]
        
        if not routes1 or not routes2:
            return parent1 if routes1 else parent2

        child_routes = []
        used_nodes = set()
        
        for route in routes1[:len(routes1)//2]:
            route_nodes = [node.id for node in route]
            if not any(node_id in used_nodes for node_id in route_nodes):
                child_routes.append(route)
                used_nodes.update(route_nodes)

        for route in routes2:
            route_nodes = [node.id for node in route]
            if not any(node_id in used_nodes for node_id in route_nodes):
                child_routes.append(route)
                used_nodes.update(route_nodes)

        child_vehicles = []
        for route in child_routes:
            vehicle = Vehicle(self.vrptw.vehicle_capacity)
            vehicle.route = [self.vrptw.depot] + list(route) + [self.vrptw.depot]
            
            vehicle.time = 0
            vehicle.load = 0
            vehicle.arrival_times = [0]
            
            for i in range(1, len(vehicle.route)):
                node = vehicle.route[i]
                prev_node = vehicle.route[i-1]
                arrival_time = vehicle.time + self.vrptw.distance(prev_node, node)
                service_start = max(arrival_time, node.early) if i < len(vehicle.route)-1 else arrival_time
                vehicle.time = service_start + (node.service_time if i < len(vehicle.route)-1 else 0)
                vehicle.arrival_times.append(round(arrival_time, 3))
                vehicle.load += node.demand if i < len(vehicle.route)-1 else 0
                
            child_vehicles.append(vehicle)

        return child_vehicles

    def mutate(self, solution: List[Vehicle]) -> List[Vehicle]:
        if random.random() < self.mutation_rate:
            mutation_type = random.choice(['swap', 'insert', 'reverse'])
            
            if mutation_type == 'swap':
                routes = [v.route[1:-1] for v in solution if len(v.route) > 2]
                if len(routes) >= 2:
                    route1, route2 = random.sample(routes, 2)
                    if route1 and route2:
                        pos1 = random.randint(0, len(route1)-1)
                        pos2 = random.randint(0, len(route2)-1)
                        route1[pos1], route2[pos2] = route2[pos2], route1[pos1]
            
            elif mutation_type == 'insert':
                routes = [v.route[1:-1] for v in solution if len(v.route) > 2]
                if routes:
                    route = random.choice(routes)
                    if len(route) > 1:
                        pos1 = random.randint(0, len(route)-1)
                        pos2 = random.randint(0, len(route)-1)
                        node = route.pop(pos1)
                        route.insert(pos2, node)
            
            else: 
                routes = [v.route[1:-1] for v in solution if len(v.route) > 2]
                if routes:
                    route = random.choice(routes)
                    if len(route) > 2:
                        pos1 = random.randint(0, len(route)-2)
                        pos2 = random.randint(pos1+1, len(route)-1)
                        route[pos1:pos2+1] = reversed(route[pos1:pos2+1])

        return solution

    def optimize(self, constructive_solution: List[Vehicle]) -> Tuple[List[Vehicle], float]:
        population = self.create_initial_population(constructive_solution)
        
        for generation in range(self.generations):
            fitnesses = [self.fitness(solution) for solution in population]
            
            min_fitness_idx = min(range(len(fitnesses)), key=fitnesses.__getitem__)
            if fitnesses[min_fitness_idx] < self.best_fitness:
                self.best_fitness = fitnesses[min_fitness_idx]
                self.best_solution = copy.deepcopy(population[min_fitness_idx])
            
            parents = self.select_parents(population, fitnesses)
            
            new_population = []
            
            sorted_indices = sorted(range(len(fitnesses)), key=lambda k: fitnesses[k])
            elite_solutions = [population[i] for i in sorted_indices[:self.elite_size]]
            new_population.extend(copy.deepcopy(elite_solutions))
            
            while len(new_population) < self.population_size:
                parent1, parent2 = random.sample(parents, 2)
                child = self.crossover(parent1, parent2)
                child = self.mutate(child)
                new_population.append(child)
            
            population = new_population
            
            if (generation + 1) % 10 == 0:
                print(f"Generación {generation + 1}/{self.generations}")
                print(f"Mejor fitness: {self.best_fitness}")
                print(f"Número de vehículos: {len([v for v in self.best_solution if len(v.route) > 2])}")
                print(f"Distancia total: {sum(self.vrptw.distance(route[i], route[i + 1]) for vehicle in self.best_solution for route in [vehicle.route] for i in range(len(route) - 1))}")
        
        return self.best_solution, self.best_fitness

def solve_instance_with_ga(instance_path):
    nodes, capacity = read_instance(instance_path)
    vrptw = VRPTW(nodes, capacity)
    
    print("Generando solución constructiva inicial...")
    constructive_solution = vrptw.construct_solution()
    
    constructive_distance = vrptw.calculate_total_distance()
    
    constructive_solution = vrptw.vehicles
    
    print(f"Solución constructiva: {len(constructive_solution)} vehículos, distancia: {constructive_distance}")
    
    print("\nAplicando algoritmo genético...")
    ga = GeneticVRPTW(vrptw)
    start_time = time.time()
    best_solution, ga_fitness = ga.optimize(constructive_solution)
    
    
    computation_time = (time.time() - start_time) * 1000
    
    total_distance = sum(
        vrptw.distance(route[i], route[i + 1])
        for vehicle in best_solution
        for route in [vehicle.route]
        for i in range(len(route) - 1)
    )
    
    routes = [(vehicle.route, vehicle.arrival_times[-1], vehicle.load) 
              for vehicle in best_solution]
    
    return best_solution, total_distance, computation_time, routes, capacity

def update_vehicle_times(self, vehicle):
    time = 0
    arrival_times = [0]  
    
    for i in range(1, len(vehicle.route)):
        current_node = vehicle.route[i-1]
        next_node = vehicle.route[i]
        
        arrival_time = time + self.distance(current_node, next_node)
        service_start = max(arrival_time, next_node.early) if i < len(vehicle.route)-1 else arrival_time
        time = service_start + (next_node.service_time if i < len(vehicle.route)-1 else 0)
        
        arrival_times.append(round(arrival_time, 3))
    
    vehicle.time = time
    vehicle.arrival_times = arrival_times

def read_instance(filename):
    with open(filename, 'r') as f:
        n, capacity = map(int, f.readline().split())
        nodes = []
        for _ in range(n + 1):
            id, x, y, demand, early, late, service_time = map(float, f.readline().split())
            nodes.append(Node(int(id), x, y, demand, early, late, service_time))
    return nodes, capacity

def save_results_to_excel(instance_name, vehicles, total_distance, computation_time, vehicle_capacity):
    if not os.path.exists(OUTPUT_FILE):
        workbook = Workbook()
        workbook.remove(workbook.active) 
    else:
        workbook = openpyxl.load_workbook(OUTPUT_FILE)

    sheet = workbook.create_sheet(title=instance_name)

    used_vehicles = [vehicle for vehicle in vehicles if len(vehicle.route) > 2]

    sheet.append([len(used_vehicles), round(total_distance, 3), round(computation_time, 3)])
    
    for vehicle in used_vehicles:
        route = [node.id for node in vehicle.route]
        adjusted_times = [max(at, node.early) for at, node in zip(vehicle.arrival_times, vehicle.route)]

        route_info = [len(route) - 2] + route + adjusted_times + [vehicle.load]
        sheet.append(route_info)

    workbook.save(OUTPUT_FILE)

def main():
    for instance_filename in os.listdir(INSTANCES_DIR):
        if instance_filename.endswith('.txt'):
            instance_path = os.path.join(INSTANCES_DIR, instance_filename)
            instance_name = instance_filename.replace('.txt', '')
            
            print(f"\nResolviendo instancia: {instance_name}")
            
            vehicles, total_distance, computation_time, routes, vehicle_capacity = (
                solve_instance_with_ga(instance_path)
            )
            
            print(f"\nResultados finales:")
            print(f"Vehículos: {len(vehicles)}")
            print(f"Distancia total: {total_distance}")
            print(f"Tiempo de cómputo: {computation_time:.2f} ms")
            print(f"Capacidad del vehículo: {vehicle_capacity}")
            print("Rutas:")
            for i, (route, arrival_time, load) in enumerate(routes, 1):
                print(f"  Ruta {i}:")
                print(f"    Nodos: {' -> '.join(str(node.id) for node in route)}")
                print(f"    Tiempo de llegada: {arrival_time:.2f}")
                print(f"    Carga: {load}")
            
            save_results_to_excel(instance_name, vehicles, total_distance, 
                                computation_time, vehicle_capacity)

if __name__ == '__main__':
    main()