import os
import openpyxl
from openpyxl import Workbook
import math
import time
import random

OUTPUT_FILE = 'VRPTW_LuisaMariaAlvarez_GRASPALPHA.xlsx'
INSTANCES_DIR = 'instances'

class Node:
    def __init__(self, id, x, y, demand, early, late, service_time):
        self.id = id
        self.x = x
        self.y = y
        self.demand = demand
        self.early = early
        self.late = late
        self.service_time = service_time

class Vehicle:
    def __init__(self, capacity):
        self.capacity = capacity
        self.route = []
        self.load = 0
        self.time = 0
        self.arrival_times = []

class VRPTW:
    def __init__(self, nodes, vehicle_capacity, alpha=0.1):
        self.nodes = nodes
        self.vehicle_capacity = vehicle_capacity
        self.depot = nodes[0]
        self.vehicles = []
        self.alpha = alpha

    def distance(self, node1, node2):
        return round(math.sqrt((node1.x - node2.x) ** 2 + (node1.y - node2.y) ** 2), 3)

    def calculate_total_distance(self):
        total_distance = 0
        for vehicle in self.vehicles:
            route = vehicle.route
            for i in range(len(route) - 1):
                total_distance += self.distance(route[i], route[i + 1])
        return round(total_distance, 3)

    def is_feasible(self, vehicle, node):
        # Verificar la capacidad del vehículo
        if vehicle.load + node.demand > vehicle.capacity:
            return False

        # Calcular el tiempo de llegada al nodo
        arrival_time = vehicle.time + self.distance(vehicle.route[-1], node)

        # Verificar si el vehículo llega después del tiempo límite del nodo
        if arrival_time > node.late:
            return False

        # Calcular el tiempo de regreso al depósito después de visitar este nodo
        return_time_to_depot = max(arrival_time, node.early) + node.service_time + self.distance(node, self.depot)

        # Verificar si es posible regresar al depósito dentro de su ventana de tiempo
        if return_time_to_depot > self.depot.late:
            return False

        return True

    def update_vehicle(self, vehicle, node):
        # Calcular el tiempo de llegada al nodo
        arrival_time = vehicle.time + self.distance(vehicle.route[-1], node)

        # Ajustar el inicio del servicio si el vehículo llega antes de la ventana de tiempo
        service_start = max(arrival_time, node.early)

        # Actualizar el tiempo del vehículo con el tiempo de servicio en el nodo
        vehicle.time = service_start + node.service_time

        # Actualizar la carga del vehículo y agregar el nodo a la ruta
        vehicle.route.append(node)
        vehicle.load += node.demand
        vehicle.arrival_times.append(round(arrival_time, 3))  # Registrar el tiempo de llegada real

        # Debugging: Imprimir información sobre el nodo agregado
        print(f"Node {node.id} added to vehicle with arrival time {arrival_time} and start of service at {service_start}")
        print(f"Current vehicle load: {vehicle.load}, current route: {[n.id for n in vehicle.route]}")
        print(f"Vehicle arrival times: {vehicle.arrival_times}\n")

    def construct_solution(self):
        unassigned = self.nodes[1:]  # Nodos no asignados (todos menos el depósito)
        max_attempts = len(self.nodes) * 2  # Número máximo de intentos
        attempts = 0

        while unassigned and attempts < max_attempts:
            vehicle = Vehicle(self.vehicle_capacity)
            vehicle.route = [self.depot]
            vehicle.time = 0
            vehicle.arrival_times = [0]
            assigned_any = False

            while True:
                feasible_nodes = [node for node in unassigned if self.is_feasible(vehicle, node)]

                if not feasible_nodes:
                    break

                # Ordenar los nodos factibles por la distancia al nodo actual
                feasible_nodes.sort(key=lambda node: self.distance(vehicle.route[-1], node))
                rcl_size = max(1, int(len(feasible_nodes) * self.alpha))  # Tamaño de la lista restringida (RCL)
                rcl = feasible_nodes[:rcl_size]  # Seleccionar la lista restringida

                next_node = random.choice(rcl)  # Seleccionar aleatoriamente de la RCL

                self.update_vehicle(vehicle, next_node)
                unassigned.remove(next_node)
                assigned_any = True

            # Completar la ruta regresando al depósito
            vehicle.route.append(self.depot)
            vehicle.arrival_times.append(round(vehicle.time + self.distance(vehicle.route[-2], self.depot), 3))
            self.vehicles.append(vehicle)

            if not assigned_any:
                attempts += 1
            else:
                attempts = 0

        return self.vehicles

    def calculate_route_distance(self, route):
        total_distance = 0
        for i in range(len(route) - 1):
            total_distance += self.distance(route[i], route[i + 1])
        return total_distance

    def print_distance_matrix(self):
        n = len(self.nodes)
        distance_matrix = [[0] * n for _ in range(n)]
        
        for i in range(n):
            for j in range(n):
                distance_matrix[i][j] = round(math.sqrt((self.nodes[i].x - self.nodes[j].x)**2 + (self.nodes[i].y - self.nodes[j].y)**2), 3)
        
        print("Matriz de Distancias:")
        for row in distance_matrix:
            print(row)

def read_instance(filename):
    with open(filename, 'r') as f:
        n, capacity = map(int, f.readline().split())
        nodes = []
        for _ in range(n + 1):
            id, x, y, demand, early, late, service_time = map(float, f.readline().split())
            nodes.append(Node(int(id), x, y, demand, early, late, service_time))
    return nodes, capacity

def save_results_to_excel(instance_name, vehicles, total_distance, computation_time, vehicle_capacity):
    if not os.path.exists(OUTPUT_FILE):
        workbook = Workbook()
        workbook.remove(workbook.active)
    else:
        workbook = openpyxl.load_workbook(OUTPUT_FILE)

    sheet = workbook.create_sheet(title=instance_name)

    used_vehicles = [vehicle for vehicle in vehicles if len(vehicle.route) > 2]

    sheet.append([len(used_vehicles), round(total_distance, 3), round(computation_time, 3)])

    for vehicle in used_vehicles:
        route = [node.id for node in vehicle.route]
        adjusted_times = [max(at, node.early) for at, node in zip(vehicle.arrival_times, vehicle.route)]

        route_info = [len(route) - 2] + route + adjusted_times + [vehicle.load]
        sheet.append(route_info)

    workbook.save(OUTPUT_FILE)

def solve_instance(instance_path):
    nodes, capacity = read_instance(instance_path)
    vrptw = VRPTW(nodes, capacity)
    vrptw.print_distance_matrix()
    start_time = time.time()
    solution = vrptw.construct_solution()
    end_time = time.time()

    computation_time = (end_time - start_time) * 1000  # Tiempo en milisegundos
    total_distance = vrptw.calculate_total_distance()

    routes = [(vehicle.route, vehicle.arrival_times[-1], vehicle.load) for vehicle in solution]

    return solution, total_distance, computation_time, routes, capacity

def grasp(instance, max_iterations=10, alpha=0.1):
    best_solution = None
    best_distance = float('inf')

    for _ in range(max_iterations):
        vrptw = VRPTW(instance.nodes, instance.vehicle_capacity, alpha)
        solution = vrptw.construct_solution()

        total_distance = vrptw.calculate_total_distance()

        if total_distance < best_distance:
            best_solution = solution
            best_distance = total_distance

    return best_solution, best_distance

def main():
    for instance_filename in os.listdir(INSTANCES_DIR):
        if instance_filename.endswith('.txt'):
            instance_path = os.path.join(INSTANCES_DIR, instance_filename)
            instance_name = instance_filename.replace('.txt', '')
            vehicles, total_distance, computation_time, routes, vehicle_capacity = solve_instance(instance_path)
            save_results_to_excel(instance_name, vehicles, total_distance, computation_time, vehicle_capacity)

if __name__ == '__main__':
    main()
