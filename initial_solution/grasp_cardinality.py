import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import math
import time
import random

OUTPUT_FILE = 'VRPTW_LuisaMariaAlvarez_GRASPCARDINALITY.xlsx'
INSTANCES_DIR = 'instances'  

class Node:
    def __init__(self, id, x, y, demand, early, late, service_time):
        self.id = id
        self.x = x
        self.y = y
        self.demand = demand
        self.early = early
        self.late = late
        self.service_time = service_time

class Vehicle:
    def __init__(self, capacity):
        self.capacity = capacity
        self.route = []
        self.load = 0
        self.time = 0
        self.arrival_times = []

class VRPTW:
    def __init__(self, nodes, vehicle_capacity):
        self.nodes = nodes
        self.vehicle_capacity = vehicle_capacity
        self.depot = nodes[0]
        self.vehicles = []
        self.RCL_SIZE = 3 

    def distance(self, node1, node2):
        return round(math.sqrt((node1.x - node2.x)**2 + (node1.y - node2.y)**2), 3)

    def calculate_total_distance(self):
        total_distance = 0
        for vehicle in self.vehicles:
            route = vehicle.route
            for i in range(len(route) - 1):
                total_distance += self.distance(route[i], route[i+1])
        return round(total_distance, 3)

    def is_feasible(self, vehicle, node):
        if vehicle.load + node.demand > vehicle.capacity:
            return False
        
        arrival_time = vehicle.time + self.distance(vehicle.route[-1], node)
        
        if arrival_time > node.late:
            return False
        return_time_to_depot = max(arrival_time, node.early) + node.service_time + self.distance(node, self.depot)

        if return_time_to_depot > self.depot.late:
            return False
        
        return True

    def update_vehicle(self, vehicle, node):
        arrival_time = vehicle.time + self.distance(vehicle.route[-1], node)
        service_start = max(arrival_time, node.early)
        vehicle.time = service_start + node.service_time
        
        print(f"Node {node.id} added to vehicle {vehicle}")
        print(f"Arrival Time: {arrival_time}")
        print(f"Service Start: {service_start}")
        print(f"Vehicle Time: {vehicle.time}")
        print(f"Vehicle Load: {vehicle.load}")
        print(f"Vehicle Route: {[node.id for node in vehicle.route]}")
        print(f"Vehicle Arrival Times: {vehicle.arrival_times}")
        print()
        
        vehicle.route.append(node)
        vehicle.load += node.demand
        vehicle.arrival_times.append(round(arrival_time, 3))

    def construct_solution(self):
        unassigned = self.nodes[1:]
        max_attempts = len(self.nodes) * 2
        attempts = 0
        
        while unassigned and attempts < max_attempts:
            vehicle = Vehicle(self.vehicle_capacity)
            vehicle.route = [self.depot]
            vehicle.time = 0
            vehicle.arrival_times = [0]
            assigned_any = False
            
            while True:
                feasible_nodes = [node for node in unassigned if self.is_feasible(vehicle, node)]
                
                if not feasible_nodes:
                    break
                
               
                feasible_nodes.sort(key=lambda node: self.distance(vehicle.route[-1], node))
                
                rcl_size = min(self.RCL_SIZE, len(feasible_nodes))
                rcl = feasible_nodes[:rcl_size]
                
                next_node = random.choice(rcl)
                
                self.update_vehicle(vehicle, next_node)
                unassigned.remove(next_node)
                assigned_any = True
            
            vehicle.route.append(self.depot)
            vehicle.arrival_times.append(round(vehicle.time + self.distance(vehicle.route[-2], self.depot), 3))
            self.vehicles.append(vehicle)
            
            if not assigned_any:
                attempts += 1
            else:
                attempts = 0
        
        return self.vehicles

def read_instance(filename):
    with open(filename, 'r') as f:
        n, capacity = map(int, f.readline().split())
        nodes = []
        for _ in range(n + 1):
            id, x, y, demand, early, late, service_time = map(float, f.readline().split())
            nodes.append(Node(int(id), x, y, demand, early, late, service_time))
    return nodes, capacity

def save_results_to_excel(instance_name, vehicles, total_distance, computation_time, vehicle_capacity):
    if not os.path.exists(OUTPUT_FILE):
        workbook = Workbook()
        workbook.remove(workbook.active) 
    else:
        workbook = openpyxl.load_workbook(OUTPUT_FILE)

    sheet = workbook.create_sheet(title=instance_name)

    used_vehicles = [vehicle for vehicle in vehicles if len(vehicle.route) > 2]

    sheet.append([len(used_vehicles), round(total_distance, 3), round(computation_time, 3)])
    
    for vehicle in used_vehicles:
        route = [node.id for node in vehicle.route]
        adjusted_times = [max(at, node.early) for at, node in zip(vehicle.arrival_times, vehicle.route)]

        route_info = [len(route) - 2] + route + adjusted_times + [vehicle.load]
        sheet.append(route_info)
        
    workbook.save(OUTPUT_FILE)


def solve_instance(instance_path):
    nodes, capacity = read_instance(instance_path)
    vrptw = VRPTW(nodes, capacity)

    start_time = time.time()
    solution = vrptw.construct_solution()
    end_time = time.time()

    computation_time = (end_time - start_time) * 1000  
    total_distance = vrptw.calculate_total_distance()

    routes = [(vehicle.route, vehicle.arrival_times[-1], vehicle.load) for vehicle in solution]

    return solution, total_distance, computation_time, routes, capacity

def main():
    for instance_filename in os.listdir(INSTANCES_DIR):
        if instance_filename.endswith('.txt'):
            instance_path = os.path.join(INSTANCES_DIR, instance_filename)
            instance_name = instance_filename.replace('.txt', '')
            vehicles, total_distance, computation_time, routes, vehicle_capacity = solve_instance(instance_path)
            save_results_to_excel(instance_name, vehicles, total_distance, computation_time, vehicle_capacity)

if __name__ == '__main__':
    main()
